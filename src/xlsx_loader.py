"""xlsx_loader.py — Parse multi-tab Excel workbooks for capability input.

Each capability has up to 2 Excel workbooks:
  - CurrentFlows.xlsx (or R3_CurrentFlows.xlsx)  → current/legacy state
  - FutureFlows.xlsx  (or R3_FutureFlows.xlsx)   → future/target state

Each workbook contains these tabs (all optional except "Flows"):
  Tab Name               │ SAD Section    │ Description
  ───────────────────────┼────────────────┼──────────────────────────────────
  Flows                  │ §5.1 / §5.2   │ 47-col Application+Data+Tech flows
  Business Drivers       │ §2.2           │ Strategic drivers and priorities
  Success Criteria       │ §2.3           │ KPIs, targets, baselines
  Business Architecture  │ §3             │ Process steps (BPMN alternative)
  NFRs                   │ §6.3           │ Non-Functional Requirements
  Security Controls      │ §6.4           │ Auth, encryption, compliance
  SAP Dev Status         │ §6.2           │ Transport counts by environment
  Recommendations        │ §7.3           │ Prioritized action items

Backward Compatibility:
  - If only CSV files exist, the existing csv_parser + context_loader logic is used
  - If an xlsx file exists, it takes precedence over individual CSVs
  - The parser gracefully handles missing tabs (returns empty lists)
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl

from .context_loader import (
    BusinessDriver,
    SuccessCriterion,
    NFRequirement,
    SecurityControl,
    SAPDevStatusRow,
    Recommendation,
    CapabilityContext,
)
from .csv_parser import FlowHop, FlowChain, FlowSet, _row_to_hop, _group_chains, _normalize_header
from .bpmn_parser import BPMNNode, BPMNFlow, BPMNProcess


# ---------------------------------------------------------------------------
# Tab name constants (case-insensitive matching)
# ---------------------------------------------------------------------------
TAB_FLOWS = "Flows"
TAB_BUSINESS_DRIVERS = "Business Drivers"
TAB_SUCCESS_CRITERIA = "Success Criteria"
TAB_BUSINESS_ARCH = "Business Architecture"
TAB_NFRS = "NFRs"
TAB_SECURITY = "Security Controls"
TAB_SAP_STATUS = "SAP Dev Status"
TAB_RECOMMENDATIONS = "Recommendations"

ALL_TAB_NAMES = [
    TAB_FLOWS,
    TAB_BUSINESS_DRIVERS,
    TAB_SUCCESS_CRITERIA,
    TAB_BUSINESS_ARCH,
    TAB_NFRS,
    TAB_SECURITY,
    TAB_SAP_STATUS,
    TAB_RECOMMENDATIONS,
]


# ---------------------------------------------------------------------------
# Generic sheet reader
# ---------------------------------------------------------------------------
def _read_sheet_as_dicts(wb: openpyxl.Workbook, tab_name: str) -> list[dict[str, str]]:
    """Read a named sheet from a workbook into a list of row dicts.

    Returns empty list if the tab doesn't exist or has no data rows.
    Uses case-insensitive tab name matching.
    """
    # Case-insensitive tab lookup
    sheet_map = {name.lower().strip(): name for name in wb.sheetnames}
    actual_name = sheet_map.get(tab_name.lower().strip())
    if not actual_name:
        return []

    ws = wb[actual_name]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return []

    headers = [_normalize_header(str(h or "")) for h in header_row]
    result = []
    for row in rows_iter:
        row_dict = {}
        for i in range(min(len(headers), len(row))):
            val = row[i]
            row_dict[headers[i]] = str(val).strip() if val is not None else ""
        # Skip completely empty rows
        if any(v for v in row_dict.values()):
            result.append(row_dict)
    return result


# ---------------------------------------------------------------------------
# Tab parsers
# ---------------------------------------------------------------------------
def _parse_flows_tab(wb: openpyxl.Workbook, label: str, source_file: str) -> FlowSet:
    """Parse the 'Flows' tab into a FlowSet."""
    rows = _read_sheet_as_dicts(wb, TAB_FLOWS)
    hops = []
    for row in rows:
        hop = _row_to_hop(row)
        if hop:
            hops.append(hop)
    chains = _group_chains(hops)
    return FlowSet(label=label, source_file=source_file, hops=hops, chains=chains)


def _parse_business_drivers_tab(wb: openpyxl.Workbook) -> list[BusinessDriver]:
    """Parse the 'Business Drivers' tab."""
    rows = _read_sheet_as_dicts(wb, TAB_BUSINESS_DRIVERS)
    drivers = []
    for row in rows:
        name = (row.get("Driver Name") or "").strip()
        if name:
            drivers.append(BusinessDriver(
                number=(row.get("Driver #") or "").strip(),
                name=name,
                description=(row.get("Description") or "").strip(),
                strategic_alignment=(row.get("Strategic Alignment") or "").strip(),
                priority=(row.get("Priority") or "").strip(),
            ))
    return drivers


def _parse_success_criteria_tab(wb: openpyxl.Workbook) -> list[SuccessCriterion]:
    """Parse the 'Success Criteria' tab."""
    rows = _read_sheet_as_dicts(wb, TAB_SUCCESS_CRITERIA)
    criteria = []
    for row in rows:
        metric = (row.get("Metric") or "").strip()
        if metric:
            criteria.append(SuccessCriterion(
                metric=metric,
                target=(row.get("Target") or "").strip(),
                measure=(row.get("Measure") or "").strip(),
                baseline=(row.get("Baseline") or "").strip(),
                owner=(row.get("Owner") or "").strip(),
            ))
    return criteria


def _parse_business_arch_tab(wb: openpyxl.Workbook, source_file: str) -> list[BPMNProcess]:
    """Parse the 'Business Architecture' tab into BPMNProcess objects.

    Groups rows by Process ID, creates synthetic BPMNNode/BPMNFlow objects
    that produce the same Mermaid output as parsed BPMN XML files.
    """
    rows = _read_sheet_as_dicts(wb, TAB_BUSINESS_ARCH)
    if not rows:
        return []

    # Group rows by Process ID
    process_groups: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        pid = (row.get("Process ID") or "").strip()
        if pid:
            process_groups.setdefault(pid, []).append(row)

    processes = []
    for pid, steps in sorted(process_groups.items()):
        # Sort steps by Step #
        steps.sort(key=lambda r: int(r.get("Step #") or "0") if (r.get("Step #") or "0").isdigit() else 0)

        # Build filename-like identifier for step_id/step_name extraction
        first_name = (steps[0].get("Step Name") or "").strip()
        file_name = f"{pid} {first_name}"

        proc = BPMNProcess(
            file_name=file_name,
            process_name=first_name,
            process_id=pid,
        )

        # Collect lanes and build nodes
        lane_set: set[str] = set()
        prev_node_id: Optional[str] = None

        for step in steps:
            step_num = (step.get("Step #") or "0").strip()
            step_name = (step.get("Step Name") or "").strip()
            step_desc = (step.get("Step Description") or "").strip()
            step_type = (step.get("Step Type") or "task").strip().lower()
            lane = (step.get("Lane / Role") or "").strip()
            gw_type = (step.get("Gateway Type") or "").strip().lower()
            gw_cond = (step.get("Gateway Condition") or "").strip()

            # Normalize step type to BPMN types
            type_map = {
                "usertask": "userTask",
                "user task": "userTask",
                "user": "userTask",
                "servicetask": "serviceTask",
                "service task": "serviceTask",
                "service": "serviceTask",
                "system": "serviceTask",
                "gateway": "exclusiveGateway",
                "exclusive": "exclusiveGateway",
                "parallel": "parallelGateway",
                "inclusive": "inclusiveGateway",
                "startevent": "startEvent",
                "start event": "startEvent",
                "start": "startEvent",
                "endevent": "endEvent",
                "end event": "endEvent",
                "end": "endEvent",
                "subprocess": "subProcess",
                "sub process": "subProcess",
                "task": "task",
            }
            bpmn_type = type_map.get(step_type, "task")

            # Override type if gateway type is specified
            if gw_type and gw_type != "none":
                gw_map = {
                    "exclusive": "exclusiveGateway",
                    "parallel": "parallelGateway",
                    "inclusive": "inclusiveGateway",
                }
                bpmn_type = gw_map.get(gw_type, "exclusiveGateway")

            node_id = f"{pid}_{step_num}".replace("-", "_").replace(" ", "_")
            display_name = step_name or step_desc or f"Step {step_num}"

            node = BPMNNode(
                id=node_id,
                name=display_name,
                type=bpmn_type,
                lane=lane,
            )
            proc.nodes[node_id] = node

            if lane:
                lane_set.add(lane)

            # Build sequence flow from preceding step
            if prev_node_id:
                flow_name = gw_cond if gw_cond else ""
                flow = BPMNFlow(
                    id=f"flow_{prev_node_id}_to_{node_id}",
                    source_ref=prev_node_id,
                    target_ref=node_id,
                    name=flow_name,
                )
                proc.flows.append(flow)

            # Also handle explicit Preceding/Following Step references
            preceding = (step.get("Preceding Step") or "").strip()
            if preceding and preceding != str(int(step_num) - 1 if step_num.isdigit() else ""):
                # Non-sequential predecessor — add additional flow
                pred_id = f"{pid}_{preceding}".replace("-", "_").replace(" ", "_")
                if pred_id in proc.nodes and pred_id != prev_node_id:
                    proc.flows.append(BPMNFlow(
                        id=f"flow_{pred_id}_to_{node_id}_alt",
                        source_ref=pred_id,
                        target_ref=node_id,
                        name=gw_cond,
                    ))

            prev_node_id = node_id

        proc.lanes = sorted(lane_set)
        processes.append(proc)

    return processes


def _parse_nfrs_tab(wb: openpyxl.Workbook) -> list[NFRequirement]:
    """Parse the 'NFRs' tab."""
    rows = _read_sheet_as_dicts(wb, TAB_NFRS)
    nfrs = []
    for row in rows:
        req = (row.get("Requirement") or "").strip()
        if req:
            nfrs.append(NFRequirement(
                category=(row.get("Category") or "").strip(),
                requirement=req,
                target_sla=(row.get("Target / SLA") or "").strip(),
                priority=(row.get("Priority") or "").strip(),
                notes=(row.get("Notes") or "").strip(),
            ))
    return nfrs


def _parse_security_tab(wb: openpyxl.Workbook) -> list[SecurityControl]:
    """Parse the 'Security Controls' tab."""
    rows = _read_sheet_as_dicts(wb, TAB_SECURITY)
    controls = []
    for row in rows:
        approach = (row.get("Approach") or "").strip()
        if approach:
            controls.append(SecurityControl(
                concern=(row.get("Concern") or "").strip(),
                approach=approach,
                standard_policy=(row.get("Standard / Policy") or "").strip(),
                owner=(row.get("Owner") or "").strip(),
                notes=(row.get("Notes") or "").strip(),
            ))
    return controls


def _parse_sap_status_tab(wb: openpyxl.Workbook) -> list[SAPDevStatusRow]:
    """Parse the 'SAP Dev Status' tab."""
    rows = _read_sheet_as_dicts(wb, TAB_SAP_STATUS)
    statuses = []
    for row in rows:
        obj_type = (row.get("Object Type") or "").strip()
        if obj_type:
            statuses.append(SAPDevStatusRow(
                object_type=obj_type,
                object_name=(row.get("Object Name") or "").strip(),
                environment=(row.get("Environment") or "").strip(),
                count=(row.get("Count") or "").strip(),
                status=(row.get("Status") or "").strip(),
                notes=(row.get("Notes") or "").strip(),
            ))
    return statuses


def _parse_recommendations_tab(wb: openpyxl.Workbook) -> list[Recommendation]:
    """Parse the 'Recommendations' tab."""
    rows = _read_sheet_as_dicts(wb, TAB_RECOMMENDATIONS)
    recs = []
    for row in rows:
        rec_text = (row.get("Recommendation") or "").strip()
        if rec_text:
            recs.append(Recommendation(
                number=(row.get("#") or "").strip(),
                category=(row.get("Category") or "").strip(),
                recommendation=rec_text,
                priority=(row.get("Priority") or "").strip(),
                owner=(row.get("Owner") or "").strip(),
                target_date=(row.get("Target Date") or "").strip(),
                status=(row.get("Status") or "").strip(),
            ))
    return recs


# ---------------------------------------------------------------------------
# Unified result
# ---------------------------------------------------------------------------
@dataclass
class WorkbookData:
    """All data parsed from a single multi-tab workbook."""
    source_file: str
    flows: FlowSet = field(default_factory=lambda: FlowSet(label="", source_file=""))
    context: CapabilityContext = field(default_factory=CapabilityContext)
    business_processes: list[BPMNProcess] = field(default_factory=list)

    @property
    def has_flows(self) -> bool:
        return bool(self.flows.hops)

    @property
    def has_context(self) -> bool:
        return (self.context.has_drivers or self.context.has_criteria or
                self.context.has_nfrs or self.context.has_security or
                self.context.has_sap_status or self.context.has_recommendations)

    @property
    def has_processes(self) -> bool:
        return bool(self.business_processes)


# ---------------------------------------------------------------------------
# Main loader
# ---------------------------------------------------------------------------
def load_workbook(xlsx_path: str, label: str = "") -> WorkbookData:
    """Load all tabs from a multi-tab Excel workbook.

    Args:
        xlsx_path: Path to the .xlsx file
        label: Display label for the flow set (e.g. "ICOST", "S/4 HANA")

    Returns:
        WorkbookData containing flows, context, and business processes
    """
    path = Path(xlsx_path)
    if not path.exists():
        return WorkbookData(source_file=str(path))

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    result = WorkbookData(source_file=str(path))

    # Tab 1: Flows (the main 47-column data)
    result.flows = _parse_flows_tab(wb, label or path.stem, str(path))

    # Tab 2-8: Supplementary context → CapabilityContext
    ctx = CapabilityContext()
    ctx.business_drivers = _parse_business_drivers_tab(wb)
    ctx.success_criteria = _parse_success_criteria_tab(wb)
    ctx.nfrs = _parse_nfrs_tab(wb)
    ctx.security_controls = _parse_security_tab(wb)
    ctx.sap_dev_status = _parse_sap_status_tab(wb)
    ctx.recommendations = _parse_recommendations_tab(wb)
    result.context = ctx

    # Tab 3 (special): Business Architecture → BPMNProcess list
    result.business_processes = _parse_business_arch_tab(wb, str(path))

    wb.close()
    return result


def find_workbook(data_dir: Path, release_id: str, flow_type: str) -> Optional[Path]:
    """Find the best workbook file for a given flow type.

    Priority: {Release}_{Type}.xlsx → {Type}.xlsx
    Where Type is 'CurrentFlows' or 'FutureFlows'.

    Returns the path if an xlsx is found, None otherwise (caller falls back to CSV).
    """
    xlsx_candidates = [
        f"{release_id}_{flow_type}.xlsx",
        f"{flow_type}.xlsx",
    ]
    for name in xlsx_candidates:
        p = data_dir / name
        if p.exists():
            return p
    return None
