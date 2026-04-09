"""generate_simple_template.py — Create the simplified 14-column FlowInput_SIMPLE.xlsx with dropdowns.

Run once:  python scripts/generate_simple_template.py
Output:    templates/FlowInput_SIMPLE.xlsx
"""

from pathlib import Path
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import yaml
except ImportError:
    yaml = None  # type: ignore

_ROOT = Path(__file__).resolve().parent.parent
_SYSTEM_MASTER = _ROOT / "config" / "system_master.yaml"


def _load_system_master_options() -> tuple[str, str]:
    """Load DB and platform dropdown CSV strings from system_master.yaml."""
    if yaml is None or not _SYSTEM_MASTER.exists():
        return (
            "SAP HANA,Oracle,SQL Server,PostgreSQL,MongoDB,Snowflake,Teradata,DB2,MySQL,Azure SQL,Other",
            "SAP HEC (On-Prem),SAP BTP (Cloud),Azure (Cloud),AWS (Cloud),On-Prem (Linux),On-Prem (Windows),Kubernetes,Other",
        )
    with open(_SYSTEM_MASTER, encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}
    dbs = ",".join(d["label"] for d in data.get("databases", []))
    plats = ",".join(p["label"] for p in data.get("platforms", []))
    return dbs or "Other", plats or "Other"


_DB_OPTIONS, _PLATFORM_OPTIONS = _load_system_master_options()

# ---------------------------------------------------------------------------
# Column definitions: (header, width, dropdown_values_or_None, example)
# ---------------------------------------------------------------------------
COLUMNS = [
    ("Flow Chain",              28, None, "PO Create to Invoice"),
    ("Hop #",                   8,  None, "1"),
    ("Source System",           22, None, "SAP ECC"),
    ("Source Lane",             20, None, "Finance"),
    ("Target System",           22, None, "SAP S/4HANA"),
    ("Target Lane",             20, None, "Finance"),
    ("Interface / Technology",  25,
     "IDoc,RFC,BAPI,REST API,OData,SOAP,SFTP,File,CPI,PI/PO,MuleSoft,Kafka,DB Link,Manual,Other",
     "IDoc"),
    ("Frequency",               16,
     "Real-time,Near Real-time,Hourly,Daily,Weekly,Monthly,On-demand,Batch",
     "Daily"),
    ("Data Description",        35, None, "Purchase order header + line items"),
    ("Source DB Platform",      22,
     _DB_OPTIONS,
     "Oracle"),
    ("Target DB Platform",      22,
     _DB_OPTIONS,
     "SAP HANA"),
    ("Source Tech Platform",    22,
     _PLATFORM_OPTIONS,
     "On-Prem (Linux)"),
    ("Target Tech Platform",    22,
     _PLATFORM_OPTIONS,
     "SAP HEC (On-Prem)"),
    ("Integration Pattern",     22,
     "Point-to-Point,Hub-Spoke,Publish-Subscribe,Batch File,API Gateway,Database Link",
     "Hub-Spoke"),
]

HEADER_FILL = PatternFill(start_color="0071C5", end_color="0071C5", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
OPTIONAL_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
OPTIONAL_FONT = Font(name="Calibri", bold=True, color="1565C0", size=11)
EXAMPLE_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
EXAMPLE_FONT = Font(name="Calibri", italic=True, color="888888", size=10)
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# Columns 12, 13, 14 (Source Tech Platform, Target Tech Platform, Integration Pattern) are optional
OPTIONAL_COLS = {12, 13, 14}  # 1-indexed


def main() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Flows"

    # Freeze top 2 rows (header + example)
    ws.freeze_panes = "A3"

    # Write headers
    for col_idx, (header, width, _, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if col_idx in OPTIONAL_COLS:
            cell.fill = OPTIONAL_FILL
            cell.font = OPTIONAL_FONT
        else:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write example row
    for col_idx, (_, _, _, example) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx, value=example)
        cell.fill = EXAMPLE_FILL
        cell.font = EXAMPLE_FONT
        cell.border = THIN_BORDER

    # Apply data validations (dropdowns)
    for col_idx, (header, _, dropdown, _) in enumerate(COLUMNS, 1):
        if dropdown:
            col_letter = get_column_letter(col_idx)
            dv = DataValidation(
                type="list",
                formula1=f'"{dropdown}"',
                allow_blank=True,
                showDropDown=False,
            )
            dv.errorTitle = header
            dv.error = f"Please select from the dropdown or type a custom value."
            dv.prompt = f"Select {header}"
            dv.promptTitle = header
            # Apply to rows 3-1002 (1000 data rows)
            dv.add(f"{col_letter}3:{col_letter}1002")
            ws.add_data_validation(dv)

    # Pre-format 1000 data rows
    for row_idx in range(3, 1003):
        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

    # --- Instructions tab ---
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 60
    ws2.column_dimensions["C"].width = 15

    instructions = [
        ("Column", "Description", "Required?"),
        ("Flow Chain", "Name of the end-to-end business flow (e.g., 'PO Create to Invoice')", "Yes"),
        ("Hop #", "Sequential number of this hop within the flow chain (1, 2, 3...)", "Yes"),
        ("Source System", "Application name where data originates (e.g., 'SAP ECC')", "Yes"),
        ("Source Lane", "Organizational grouping / swim lane (e.g., 'Finance', 'Procurement')", "Yes"),
        ("Target System", "Application name where data flows to", "Yes"),
        ("Target Lane", "Target organizational grouping / swim lane", "Yes"),
        ("Interface / Technology", "How the systems connect (IDoc, RFC, REST API, File, etc.)", "Yes"),
        ("Frequency", "How often the data flows (Real-time, Daily, Batch, etc.)", "Yes"),
        ("Data Description", "What data moves between systems", "Yes"),
        ("Source DB Platform", "Database hosting the source system (Oracle, HANA, SQL Server, etc.)", "Yes"),
        ("Target DB Platform", "Database hosting the target system", "Yes"),
        ("Source Tech Platform", "Hosting platform (On-Premise, Cloud, etc.) — auto-filled if blank", "Optional"),
        ("Target Tech Platform", "Hosting platform — auto-filled if blank", "Optional"),
        ("Integration Pattern", "Transport topology (Hub-Spoke, Point-to-Point, etc.) — auto-filled if blank", "Optional"),
        ("", "", ""),
        ("AUTO-ENRICHED", "The enrichment script (enrich_flows.py) will auto-populate:", ""),
        ("", "• IAPM URL, Product Owner, Business Owner (from IAPM lookup)", ""),
        ("", "• Middleware / Platform (inferred from Interface)", ""),
        ("", "• Protocol, Auth Method (inferred from Interface)", ""),
        ("", "• Data Classification (default: Intel Confidential)", ""),
        ("", "• Direction (default: →)", ""),
        ("", "• Environment Scope (default: DEV,QAS,PRD)", ""),
        ("", "• All 33 remaining columns for the full pipeline", ""),
    ]

    for row_idx, (a, b, c) in enumerate(instructions, 1):
        ws2.cell(row=row_idx, column=1, value=a).font = Font(bold=(row_idx == 1 or a == "AUTO-ENRICHED"))
        ws2.cell(row=row_idx, column=2, value=b)
        cell_c = ws2.cell(row=row_idx, column=3, value=c)
        if c == "Optional":
            cell_c.font = Font(color="1565C0", italic=True)

    # Header row styling for instructions
    for col in range(1, 4):
        ws2.cell(row=1, column=col).fill = HEADER_FILL
        ws2.cell(row=1, column=col).font = HEADER_FONT

    out_path = _ROOT / "templates" / "FlowInput_SIMPLE.xlsx"
    wb.save(out_path)
    print(f"  ✓ Created {out_path}")
    print(f"    14 columns (11 required + 3 optional)")
    print(f"    Dropdowns on: Interface, Frequency, DB platforms, Tech platforms, Integration Pattern")


if __name__ == "__main__":
    main()
