"""generate_system_registry.py — Extract IAPM application list for Input Portal dropdowns.

Produces src/data/systemRegistry.ts in the Input Portal repo with:
  - KNOWN_SYSTEMS: systems already used in existing flow files (pinned at top)
  - IAPM_SYSTEMS: all active IAPM apps (acronym + name) for autocomplete
  - DB_OPTIONS / PLATFORM_OPTIONS: locked dropdown values from system_master.yaml
  - SYSTEM_DEFAULTS: auto-fill DB + Platform when a system is selected

Usage:
    python scripts/generate_system_registry.py [--portal-dir <path>]
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from pathlib import Path

import yaml

try:
    import openpyxl
except ImportError:
    openpyxl = None  # type: ignore

_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_PORTAL = Path(r"C:\Users\sajivfra\Documents\IAO-Architecture-Input-Portal")
SYSTEM_MASTER = _ROOT / "config" / "system_master.yaml"


def _extract_flow_systems(towers_dir: Path) -> set[str]:
    """Extract unique system names from all existing flow XLSX files."""
    systems: set[str] = set()
    if not openpyxl:
        return systems

    for root, _, files in os.walk(towers_dir):
        for fn in files:
            if not fn.endswith("Flows.xlsx"):
                continue
            try:
                wb = openpyxl.load_workbook(
                    os.path.join(root, fn), read_only=True, data_only=True
                )
                ws = wb["Flows"] if "Flows" in wb.sheetnames else wb.active
                headers = [str(c.value or "") for c in next(ws.iter_rows(max_row=1))]
                src_idx = next(
                    (i for i, h in enumerate(headers) if "Source System" in h), None
                )
                tgt_idx = next(
                    (i for i, h in enumerate(headers) if "Target System" in h), None
                )
                for row in ws.iter_rows(min_row=2, values_only=True):
                    for idx in (src_idx, tgt_idx):
                        if idx is not None and row[idx]:
                            val = str(row[idx]).strip()
                            if val and not val.startswith("e.g."):
                                systems.add(val)
                wb.close()
            except Exception:
                pass
    return systems


def _extract_iapm_apps(iapm_csv: Path) -> list[dict[str, str]]:
    """Extract active IAPM applications (Deployed/Developing/Planning)."""
    apps: list[dict[str, str]] = []
    seen: set[str] = set()

    with open(iapm_csv, "r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            status = (row.get("applicationLifecycleStatusNm") or "").strip()
            if status not in ("Deployed", "Developing", "Planning", "Registration", "Deployed-Legal Hold"):
                continue
            acronym = (row.get("applicationAcronymNm") or "").strip()
            name = (row.get("applicationNm") or "").strip()
            app_id = (row.get("applicationId") or "").strip()
            if not acronym or acronym.lower() in seen:
                continue
            seen.add(acronym.lower())
            apps.append({
                "id": app_id,
                "acronym": acronym,
                "name": name,
                "status": status,
            })

    return sorted(apps, key=lambda a: a["acronym"].lower())


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate system registry for Input Portal")
    parser.add_argument(
        "--portal-dir",
        type=Path,
        default=DEFAULT_PORTAL,
        help="Path to IAO-Architecture-Input-Portal repo",
    )
    args = parser.parse_args()

    iapm_csv = _ROOT / "data" / "iapm" / "IAPM_All_Solutions.csv"
    towers_dir = _ROOT / "towers"

    # 1. Extract known systems from flow files
    known = sorted(_extract_flow_systems(towers_dir))
    print(f"  ✓ Found {len(known)} systems in existing flow files")

    # 2. Extract all active IAPM apps
    iapm_apps = _extract_iapm_apps(iapm_csv)
    print(f"  ✓ Found {len(iapm_apps)} active IAPM applications")

    # 3. Build the combined list: known systems first, then IAPM acronyms
    # Known systems stay as-is (exact names from existing flows)
    # IAPM apps provide "acronym — full name" for discoverability
    iapm_labels = [f"{a['acronym']}" for a in iapm_apps]

    # Merge: known first, then IAPM apps not already in known
    known_lower = {s.lower() for s in known}
    extra = [label for label in iapm_labels if label.lower() not in known_lower]

    # 4. Load system_master.yaml for DB/Platform options and system defaults
    sm_data: dict = {}
    if SYSTEM_MASTER.exists():
        with open(SYSTEM_MASTER, encoding="utf-8") as f:
            sm_data = yaml.safe_load(f) or {}

    db_options = [d["label"] for d in sm_data.get("databases", [])]
    platform_options = [p["label"] for p in sm_data.get("platforms", [])]
    sys_defaults = sm_data.get("system_defaults", {})
    print(f"  ✓ Loaded system_master.yaml: {len(db_options)} DBs, "
          f"{len(platform_options)} platforms, {len(sys_defaults)} system defaults")

    # 5. Write TypeScript file
    out_path = args.portal_dir / "src" / "data" / "systemRegistry.ts"

    known_json = json.dumps(known, indent=2)
    all_json = json.dumps(known + extra, indent=2)
    db_json = json.dumps(db_options, indent=2)
    plat_json = json.dumps(platform_options, indent=2)
    defaults_json = json.dumps(sys_defaults, indent=2)

    ts_content = f"""/**
 * System registry for Source/Target System dropdowns.
 * Auto-generated by scripts/generate_system_registry.py — DO NOT EDIT.
 *
 * KNOWN_SYSTEMS: {len(known)} systems from existing flow files (pinned at top)
 * ALL_SYSTEMS:   {len(known) + len(extra)} total (known + {len(extra)} active IAPM apps)
 * DB_OPTIONS:    {len(db_options)} approved database platforms
 * PLATFORM_OPTIONS: {len(platform_options)} approved tech platforms
 * SYSTEM_DEFAULTS: {len(sys_defaults)} system -> DB/Platform auto-fill mappings
 */

/** Systems already used in existing architecture flow files. */
export const KNOWN_SYSTEMS: string[] = {known_json};

/** All selectable systems: known systems first, then remaining active IAPM apps. */
export const ALL_SYSTEMS: string[] = {all_json};

/** Approved database platform options (locked dropdown). */
export const DB_OPTIONS: string[] = {db_json};

/** Approved technology platform options (locked dropdown). */
export const PLATFORM_OPTIONS: string[] = {plat_json};

/** System -> default DB/Platform mapping for auto-fill on system selection. */
export const SYSTEM_DEFAULTS: Record<string, {{ db: string; platform: string }}> = {defaults_json};
"""

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(ts_content, encoding="utf-8")
    print(f"  ✓ Wrote {out_path}")
    print(f"    {len(known)} known + {len(extra)} IAPM = {len(known) + len(extra)} total systems")
    print(f"    {len(db_options)} DB options, {len(platform_options)} platform options, "
          f"{len(sys_defaults)} system defaults")


if __name__ == "__main__":
    main()
