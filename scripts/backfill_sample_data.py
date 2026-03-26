"""backfill_sample_data.py — Pre-fill supplementary tabs with grounded sample data.

Reads each capability's existing Flows tab + tower.yaml metadata to generate
contextually accurate sample data for:
  - Business Drivers       (§2.2)
  - Success Criteria       (§2.3)
  - NFRs                   (§6.3)
  - Security Controls      (§6.4)
  - SAP Dev Status         (§6.2)
  - Recommendations        (§7.3)

Data is grounded in the actual systems, flows, and integration patterns found
in each capability's xlsx, not generic placeholder text.

Usage:
    python backfill_sample_data.py                          # all towers
    python backfill_sample_data.py --tower FPR              # single tower
    python backfill_sample_data.py --tower FPR --cap DS-020 # single capability
    python backfill_sample_data.py --dry-run                # preview only
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl
import yaml
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WORKSPACE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(WORKSPACE))
os.chdir(str(WORKSPACE))

TOWERS_DIR = WORKSPACE / "towers"

# ---------------------------------------------------------------------------
# Styling
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(
    bottom=Side(style="thin", color="000000"),
    right=Side(style="thin", color="D9D9D9"),
)

# ---------------------------------------------------------------------------
# Tower domain knowledge — grounded in IAO program context
# ---------------------------------------------------------------------------
TOWER_CONTEXT = {
    "FPR": {
        "domain": "Finance",
        "full_name": "Finance Plan To Report",
        "strategic_goals": [
            ("S/4 HANA Finance Consolidation", "Migrate legacy costing and reporting platforms to unified S/4 HANA finance backbone", "IDM 2.0 Core Finance Transformation", "High"),
            ("Real-Time Financial Visibility", "Enable real-time cost reporting and variance analysis replacing batch-driven legacy processes", "CFO Digital Finance Initiative", "High"),
            ("Regulatory Compliance Readiness", "Ensure SOX compliance and audit trail continuity through the ERP transition period", "Intel Corporate Compliance", "Medium"),
        ],
        "success_metrics": [
            ("Month-End Close Cycle Time", "< 3 business days", "Calendar days from period close trigger to final posting", "5 business days (legacy)", "Finance Controller"),
            ("Cost Variance Accuracy", "< 0.5% deviation", "Variance between standard and actual cost post-migration", "1.2% (ICOST baseline)", "Cost Accounting Lead"),
            ("System Availability (Finance)", "99.9% uptime", "S/4 HANA finance module availability during business hours", "99.5% (legacy)", "IT Operations"),
        ],
        "nfr_profile": "financial_batch",
    },
    "FTS-IF": {
        "domain": "Supply Chain (Intel Foundry)",
        "full_name": "Forecast to Stock - Intel Foundry",
        "strategic_goals": [
            ("Intel Foundry Supply Chain Integration", "Integrate Intel Foundry manufacturing and logistics into unified S/4 HANA supply chain", "IDM 2.0 Foundry Enablement", "High"),
            ("Warehouse & Logistics Modernization", "Modernize warehouse management and shipping processes with EWM integration", "Supply Chain Digital Transformation", "High"),
            ("Production Planning Optimization", "Enable MRP-driven production planning with real-time material availability", "Manufacturing Excellence", "Medium"),
        ],
        "success_metrics": [
            ("Order Fulfillment Lead Time", "< 48 hours", "Time from production completion to shipment dispatch", "72 hours (legacy)", "Logistics Manager"),
            ("Inventory Accuracy", "> 99.5%", "Physical vs system inventory match rate", "97.8% (current)", "Warehouse Manager"),
            ("MRP Planning Cycle", "< 4 hours", "End-to-end MRP run including exception processing", "8 hours (legacy)", "Planning Lead"),
        ],
        "nfr_profile": "manufacturing_realtime",
    },
    "FTS-IP": {
        "domain": "Supply Chain (Intel Products)",
        "full_name": "Forecast to Stock - Intel Products",
        "strategic_goals": [
            ("Intel Products Supply Chain Unification", "Consolidate Intel Products manufacturing and logistics onto S/4 HANA platform", "IDM 2.0 Products Transformation", "High"),
            ("End-to-End Traceability", "Enable lot/batch traceability from raw material to finished goods shipment", "Quality & Compliance", "High"),
            ("Demand-Supply Matching", "Implement responsive demand and supply matching (RDSM) for IP product lines", "Supply Chain Agility", "Medium"),
        ],
        "success_metrics": [
            ("Production Schedule Adherence", "> 95%", "Percentage of production orders completed on schedule", "88% (current)", "Production Manager"),
            ("Material Availability Rate", "> 98%", "Materials available at point of need for production", "94% (current)", "Materials Planning"),
            ("Shipping On-Time Delivery", "> 97%", "Orders shipped within committed delivery window", "93% (current)", "Logistics Lead"),
        ],
        "nfr_profile": "manufacturing_realtime",
    },
    "OTC-IF": {
        "domain": "Order Management (Intel Foundry)",
        "full_name": "Order To Cash - Intel Foundry",
        "strategic_goals": [
            ("Foundry Customer Order Digitization", "Digitize end-to-end order capture, pricing, and fulfillment for Intel Foundry customers", "IDM 2.0 Foundry Revenue", "High"),
            ("Global Trade Compliance Automation", "Automate export/import compliance screening and customs declarations", "Global Trade Operations", "High"),
            ("Revenue Recognition Accuracy", "Ensure compliant revenue recognition aligned with ASC 606 through S/4 HANA billing", "Finance & Compliance", "Medium"),
        ],
        "success_metrics": [
            ("Order-to-Cash Cycle Time", "< 5 business days", "End-to-end cycle from order capture to cash application", "8 business days (legacy)", "OTC Process Owner"),
            ("Trade Compliance Screening Rate", "100%", "Orders screened for denied parties and export controls", "99.2% (current)", "Global Trade Manager"),
            ("Billing Accuracy", "> 99.8%", "Invoices generated without errors requiring credit/re-bill", "98.5% (current)", "Billing Manager"),
        ],
        "nfr_profile": "transactional_highvol",
    },
    "OTC-IP": {
        "domain": "Order Management (Intel Products)",
        "full_name": "Order To Cash - Intel Products",
        "strategic_goals": [
            ("IP Order Management Transformation", "Transform Intel Products order management onto S/4 HANA with integrated pricing and ATP", "IDM 2.0 Products Revenue", "High"),
            ("Customer Experience Improvement", "Reduce order processing time and improve order visibility for IP customers", "Customer Centricity", "High"),
            ("Returns & Rebate Automation", "Automate returns processing, rebate management, and chargeback handling", "Revenue Assurance", "Medium"),
        ],
        "success_metrics": [
            ("Order Processing Time", "< 2 hours", "Time from order receipt to order confirmation", "6 hours (current)", "Order Management Lead"),
            ("Customer Credit Decision Time", "< 15 minutes", "Automated credit check and approval for standard orders", "2 hours (manual)", "Credit Manager"),
            ("Returns Processing Cycle", "< 3 business days", "End-to-end returns receipt to credit memo issuance", "7 business days (current)", "Returns Manager"),
        ],
        "nfr_profile": "transactional_highvol",
    },
    "PTP": {
        "domain": "Procurement",
        "full_name": "Procure To Pay",
        "strategic_goals": [
            ("Procurement Process Standardization", "Standardize procurement processes across direct, indirect, and services on S/4 HANA + Ariba", "IDM 2.0 Procurement Excellence", "High"),
            ("Supplier Collaboration Enhancement", "Enable digital supplier collaboration for consignment, subcontracting, and quality management", "Supplier Ecosystem", "High"),
            ("Payment Automation", "Automate invoice verification, three-way matching, and payment execution", "Finance Efficiency", "Medium"),
        ],
        "success_metrics": [
            ("PO Cycle Time", "< 24 hours", "Requisition approval to PO dispatch to supplier", "48 hours (current)", "Procurement Lead"),
            ("Invoice Automation Rate", "> 80%", "Invoices processed without manual intervention (touchless)", "45% (current)", "AP Manager"),
            ("Supplier On-Time Delivery", "> 95%", "Supplier adherence to confirmed delivery date", "89% (current)", "Supplier Management"),
        ],
        "nfr_profile": "transactional_highvol",
    },
    "MDM": {
        "domain": "Master Data",
        "full_name": "Master Data Management",
        "strategic_goals": [
            ("Master Data Governance", "Establish single source of truth for vendor, customer, material, and BOM master data", "IDM 2.0 Data Foundation", "High"),
            ("Data Quality Improvement", "Implement data quality rules and automated validation for master data creation and changes", "Data Governance", "High"),
            ("Cross-System Data Synchronization", "Ensure master data consistency across S/4 HANA, MDG, and downstream systems", "Enterprise Integration", "Medium"),
        ],
        "success_metrics": [
            ("Master Data Accuracy", "> 99%", "Master data records passing automated quality validation rules", "94% (current)", "Data Governance Lead"),
            ("Data Creation Lead Time", "< 4 hours", "Vendor/customer/material master data creation cycle time", "2 business days (current)", "MDM Operations"),
            ("Data Sync Latency", "< 15 minutes", "Time for master data changes to propagate to all consuming systems", "4 hours (batch)", "Integration Lead"),
        ],
        "nfr_profile": "master_data",
    },
    "E2E": {
        "domain": "Cross-Functional / End-to-End",
        "full_name": "End-to-End Integrated Processes",
        "strategic_goals": [
            ("End-to-End Process Integration", "Enable cross-tower integrated processes spanning procurement, manufacturing, and fulfillment", "IDM 2.0 Process Excellence", "High"),
            ("Intel Foundry Business Enablement", "Stand up foundry-specific business processes for external customer engagement", "Intel Foundry Services", "High"),
            ("Process Visibility & Monitoring", "Provide end-to-end process visibility across tower boundaries with integrated monitoring", "Operational Excellence", "Medium"),
        ],
        "success_metrics": [
            ("E2E Process Cycle Time", "Per process SLA", "End-to-end transaction completion within defined SLA per process", "Varies by process", "E2E Process Owner"),
            ("Cross-Tower Integration Success", "> 99%", "Transactions completing across tower boundaries without manual intervention", "92% (current)", "Integration Lead"),
            ("Process Exception Rate", "< 2%", "Transactions requiring manual exception handling", "8% (current)", "Operations Manager"),
        ],
        "nfr_profile": "transactional_highvol",
    },
}

# NFR profiles — standard non-functional requirements by workload type
NFR_PROFILES = {
    "financial_batch": [
        ("Performance", "Month-end batch costing/closing completes within SLA window", "< 4 hours end-to-end batch window", "High", "Measured at PRD; includes parallel jobs"),
        ("Availability", "S/4 HANA finance modules available during business hours", "99.9% (Mon-Fri 06:00-22:00 PST)", "High", "Excludes planned maintenance windows"),
        ("Scalability", "Support 2x transaction volume growth over 3-year horizon", "Handle 500K+ journal entries/day", "Medium", "Aligns with Intel Foundry growth projections"),
        ("Recoverability", "RPO/RTO for financial systems meets audit requirements", "RPO < 1 hour, RTO < 4 hours", "High", "SOX-critical systems"),
        ("Data Volume", "Support growing data volumes from legacy migration + BAU", "50M+ records in material ledger", "Medium", "Archiving strategy required post go-live"),
        ("Latency", "Near-real-time posting for financial transactions", "< 5 seconds for online postings", "Medium", "Batch processes have separate SLAs"),
        ("Concurrency", "Support concurrent month-end users across time zones", "200+ concurrent finance users", "Medium", "Peak during close periods"),
    ],
    "manufacturing_realtime": [
        ("Performance", "MRP/production planning run completes within defined window", "< 4 hours full MRP run", "High", "Includes exception message processing"),
        ("Availability", "Manufacturing execution systems available 24/7", "99.95% (24x7 operations)", "High", "Factory operations are continuous"),
        ("Scalability", "Support production volume increases from new product lines", "Handle 10K+ production orders/day", "High", "Intel Foundry ramp-up scenarios"),
        ("Recoverability", "Production systems recover within shift change window", "RPO < 15 min, RTO < 2 hours", "High", "Manufacturing cannot tolerate extended outages"),
        ("Data Volume", "Support high-frequency material movement transactions", "100K+ material documents/day", "Medium", "Goods receipt, issue, transfer postings"),
        ("Latency", "Real-time inventory visibility for warehouse operations", "< 2 seconds for RF/scanner transactions", "High", "EWM warehouse operations"),
        ("Concurrency", "Support factory floor workers across multiple shifts/sites", "500+ concurrent warehouse users", "Medium", "Global factory footprint"),
    ],
    "transactional_highvol": [
        ("Performance", "Order/transaction processing within interactive SLA", "< 3 seconds for online transactions", "High", "Customer-facing operations"),
        ("Availability", "Business-critical systems available during extended hours", "99.9% (06:00-22:00 all time zones)", "High", "Covers APAC, EMEA, Americas"),
        ("Scalability", "Support seasonal and promotional volume spikes", "Handle 2x baseline transaction volume", "Medium", "Quarter-end peaks"),
        ("Recoverability", "Customer-facing systems recover within business impact window", "RPO < 30 min, RTO < 2 hours", "High", "Revenue-impacting systems"),
        ("Data Volume", "Support transactional data growth from business expansion", "10M+ documents/year", "Medium", "Order, delivery, billing documents"),
        ("Latency", "Near-real-time integration for order status updates", "< 30 seconds for status propagation", "Medium", "Customer visibility requirements"),
        ("Concurrency", "Support global user base across business functions", "300+ concurrent users", "Medium", "Order management, logistics, billing"),
    ],
    "master_data": [
        ("Performance", "Master data creation and validation within user expectation", "< 5 seconds for save/validate", "High", "Includes duplicate check and validation rules"),
        ("Availability", "MDM systems available during global business hours", "99.9% (follows-the-sun)", "High", "Master data changes block downstream processes"),
        ("Scalability", "Support master data volume growth from new entities", "Handle 5M+ master records", "Medium", "Vendor + customer + material + BOM"),
        ("Recoverability", "Master data systems recover with data integrity guaranteed", "RPO = 0 (no data loss), RTO < 2 hours", "High", "Master data integrity is non-negotiable"),
        ("Data Volume", "Support comprehensive master data model across domains", "500K+ active material masters", "Medium", "Including variants and BOMs"),
        ("Latency", "Master data changes propagate to consuming systems promptly", "< 15 minutes replication latency", "High", "Blocking for transactional processing"),
        ("Concurrency", "Support centralized MDM team plus distributed maintainers", "100+ concurrent MDM users", "Medium", "Governance workflow approvals"),
    ],
}

# Standard security controls applicable to all towers (Intel enterprise)
STANDARD_SECURITY = [
    ("Authentication", "Single Sign-On (SSO) via Intel corporate Azure AD identity", "Intel IT Security Policy - Identity Management", "IT Security"),
    ("Authorization", "Role-based access control (RBAC) with SAP authorization objects", "Intel SAP Security Standards - Role Design", "SAP Security Team"),
    ("Data Classification", "All financial/operational data classified per Intel Data Classification Standard", "Intel Data Classification Policy", "Data Governance"),
    ("Data Encryption (at rest)", "AES-256 encryption for SAP HANA database and file storage", "Intel Encryption Standard", "Infrastructure Security"),
    ("Data Encryption (in transit)", "TLS 1.3 for all system-to-system and user-to-system communication", "Intel Network Security Policy", "Network Engineering"),
    ("Network Segmentation", "SAP systems in dedicated network zones with firewall controls", "Intel Network Architecture Standard", "Network Security"),
    ("API Security", "OAuth 2.0 / certificate-based authentication for all API integrations", "Intel API Security Guidelines", "Integration Architecture"),
    ("Audit Logging", "Comprehensive audit trail for all data changes and user actions (SAP Security Audit Log)", "SOX Compliance / Intel Audit Policy", "Internal Audit"),
    ("Certificate Management", "Automated certificate lifecycle management for system-to-system trust", "Intel PKI Standard", "Certificate Authority Team"),
    ("Compliance", "SOX controls, export control (EAR/ITAR) screening, data privacy (GDPR)", "Intel Corporate Compliance Framework", "Compliance Office"),
]

# Standard SAP Dev Status template
STANDARD_SAP_STATUS = [
    ("Transport Requests", "", "DEV", "—", "", "To be populated post-development"),
    ("Transport Requests", "", "QAS", "—", "", ""),
    ("Transport Requests", "", "PRD", "—", "", ""),
    ("Custom Code Objects", "", "DEV", "—", "", "CDS Views, ABAP Classes, etc."),
    ("Custom Code Objects", "", "QAS", "—", "", ""),
    ("Custom Code Objects", "", "PRD", "—", "", ""),
    ("CDS Views", "", "DEV", "—", "", "Analytical and transactional CDS"),
    ("CDS Views", "", "QAS", "—", "", ""),
    ("CDS Views", "", "PRD", "—", "", ""),
    ("Fiori Apps", "", "DEV", "—", "", "Custom + extended standard apps"),
    ("Fiori Apps", "", "QAS", "—", "", ""),
    ("Fiori Apps", "", "PRD", "—", "", ""),
    ("BAdIs / Enhancements", "", "DEV", "—", "", ""),
    ("BAdIs / Enhancements", "", "QAS", "—", "", ""),
    ("BAdIs / Enhancements", "", "PRD", "—", "", ""),
]


# ---------------------------------------------------------------------------
# Extract context from existing Flows tab
# ---------------------------------------------------------------------------
def extract_flow_context(xlsx_path: Path) -> dict:
    """Read the Flows tab and extract systems, integrations, and owners."""
    ctx = {
        "systems": set(),
        "source_lanes": set(),
        "target_lanes": set(),
        "interfaces": set(),
        "flow_chains": [],
        "owners": set(),
        "data_descriptions": [],
        "hop_count": 0,
        "chain_count": 0,
    }

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception:
        return ctx

    sheet_map = {n.lower().strip(): n for n in wb.sheetnames}
    actual = sheet_map.get("flows")
    if not actual:
        wb.close()
        return ctx

    ws = wb[actual]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return ctx

    headers = [str(h or "").strip() for h in rows[0]]
    col_idx = {h: i for i, h in enumerate(headers)}

    chains_seen = set()
    for row in rows[1:]:
        if not row or not any(str(c or "").strip() for c in row):
            continue
        ctx["hop_count"] += 1

        def val(col_name):
            i = col_idx.get(col_name, -1)
            if i >= 0 and i < len(row) and row[i]:
                return str(row[i]).strip()
            return ""

        src = val("Source System")
        tgt = val("Target System")
        if src:
            ctx["systems"].add(src)
        if tgt:
            ctx["systems"].add(tgt)

        sl = val("Source Lane")
        tl = val("Target Lane")
        if sl:
            ctx["source_lanes"].add(sl)
        if tl:
            ctx["target_lanes"].add(tl)

        iface = val("Interface / Technology")
        if iface:
            ctx["interfaces"].add(iface)

        chain = val("Flow Chain")
        if chain and chain not in chains_seen:
            chains_seen.add(chain)
            ctx["flow_chains"].append(chain)

        owner = val("Process/System Owner")
        if owner:
            ctx["owners"].add(owner)

        desc = val("Data Description")
        if desc:
            ctx["data_descriptions"].append(desc)

    ctx["chain_count"] = len(chains_seen)
    return ctx


# ---------------------------------------------------------------------------
# Generate recommendations from flow context
# ---------------------------------------------------------------------------
def generate_recommendations(tower_code: str, cap_id: str, cap_name: str,
                              flow_ctx: dict) -> list[list[str]]:
    """Generate contextual recommendations based on the capability's flow data."""
    recs = []
    num = 1

    systems = sorted(flow_ctx.get("systems", set()))
    interfaces = sorted(flow_ctx.get("interfaces", set()))
    sys_count = len(systems)
    chain_count = flow_ctx.get("chain_count", 0)

    # Architecture recommendation
    recs.append([str(num), "Architecture",
                 "Complete extended flow attributes (Data Entity, Integration Pattern, Tech Platform) in Flows tab for full BDAT coverage",
                 "High", "Tower Architect", "2026-Q2", "Open"])
    num += 1

    # Integration recommendation if many systems
    if sys_count > 5:
        top_systems = ", ".join(systems[:5])
        recs.append([str(num), "Integration",
                     "Validate integration patterns and middleware for high-complexity flows across " + str(sys_count) + " systems (" + top_systems + "...)",
                     "High", "Integration Architect", "2026-Q2", "Open"])
        num += 1

    # Data recommendation
    recs.append([str(num), "Data",
                 "Define data ownership and classification for all " + str(chain_count) + " flow chains to satisfy Data Architecture (TOGAF D) requirements",
                 "Medium", "Data Architect", "2026-Q3", "Open"])
    num += 1

    # Testing recommendation
    recs.append([str(num), "Testing",
                 "Develop integration test scenarios covering all " + str(chain_count) + " flow chains for FUT/SIT readiness",
                 "High", "Test Lead", "2026-Q3", "Open"])
    num += 1

    # Business process recommendation
    recs.append([str(num), "Business Architecture",
                 "Review and validate Business Architecture process steps against latest Signavio/BIC process models",
                 "Medium", "Business Analyst", "2026-Q2", "Open"])
    num += 1

    # Security recommendation
    recs.append([str(num), "Security",
                 "Complete security review for API integrations and data flows per Intel Security Architecture standards",
                 "Medium", "Security Architect", "2026-Q3", "Open"])
    num += 1

    return recs


# ---------------------------------------------------------------------------
# Tab definitions for writing
# ---------------------------------------------------------------------------
TAB_DEFS = {
    "Business Drivers": [
        ("Driver #", 10), ("Driver Name", 25), ("Description", 45),
        ("Strategic Alignment", 30), ("Priority", 12),
    ],
    "Success Criteria": [
        ("Metric", 25), ("Target", 20), ("Measure", 25),
        ("Baseline", 20), ("Owner", 20),
    ],
    "NFRs": [
        ("Category", 18), ("Requirement", 40), ("Target / SLA", 20),
        ("Priority", 12), ("Notes", 30),
    ],
    "Security Controls": [
        ("Concern", 20), ("Approach", 35), ("Standard / Policy", 25),
        ("Owner", 20), ("Notes", 30),
    ],
    "SAP Dev Status": [
        ("Object Type", 22), ("Object Name", 25), ("Environment", 14),
        ("Count", 10), ("Status", 14), ("Notes", 30),
    ],
    "Recommendations": [
        ("#", 6), ("Category", 18), ("Recommendation", 45),
        ("Priority", 12), ("Owner", 20), ("Target Date", 14), ("Status", 14),
    ],
}


def write_tab(wb: openpyxl.Workbook, tab_name: str, columns: list, data_rows: list) -> None:
    """Write or replace a tab in the workbook with formatted data."""
    if tab_name in wb.sheetnames:
        del wb[tab_name]

    # Determine insert position
    tab_order = ["Flows", "Business Drivers", "Success Criteria",
                 "Business Architecture", "NFRs", "Security Controls",
                 "SAP Dev Status", "Recommendations"]
    try:
        idx = tab_order.index(tab_name)
    except ValueError:
        idx = len(wb.sheetnames)
    # Clamp to number of existing sheets
    idx = min(idx, len(wb.sheetnames))

    ws = wb.create_sheet(title=tab_name, index=idx)

    # Header row
    for col_idx, (col_name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            if val and col_idx <= len(columns):
                ws.cell(row=row_idx, column=col_idx, value=str(val))

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}1"


# ---------------------------------------------------------------------------
# Main backfill logic per capability
# ---------------------------------------------------------------------------
def backfill_workbook(xlsx_path: Path, tower_code: str, cap_id: str,
                       cap_name: str, dry_run: bool = False) -> str:
    """Backfill all supplementary tabs in one workbook with grounded sample data."""
    if not xlsx_path.exists():
        return "skip-no-file"

    tower_ctx = TOWER_CONTEXT.get(tower_code, TOWER_CONTEXT.get("E2E", {}))

    # Extract context from Flows tab
    flow_ctx = extract_flow_context(xlsx_path)

    if dry_run:
        return "would-backfill-6-tabs (" + str(flow_ctx['hop_count']) + " hops, " + str(len(flow_ctx['systems'])) + " systems)"

    wb = openpyxl.load_workbook(str(xlsx_path))

    # 1. Business Drivers
    drivers = []
    for i, (name, desc, alignment, priority) in enumerate(tower_ctx.get("strategic_goals", []), 1):
        drivers.append([str(i), name, desc, alignment, priority])
    # Add capability-specific driver
    drivers.append([str(len(drivers) + 1),
                    cap_id + " Process Migration",
                    "Migrate " + cap_name + " business processes and " + str(len(flow_ctx['systems'])) + " integrated systems from legacy to S/4 HANA target architecture",
                    "IDM 2.0 " + tower_ctx.get("domain", "Transformation"),
                    "High"])
    write_tab(wb, "Business Drivers", TAB_DEFS["Business Drivers"], drivers)

    # 2. Success Criteria
    criteria = []
    for metric, target, measure, baseline, owner in tower_ctx.get("success_metrics", []):
        criteria.append([metric, target, measure, baseline, owner])
    # Add capability-specific metric
    criteria.append([
        cap_id + " Migration Completeness",
        "100% flow chains validated",
        "All " + str(flow_ctx['chain_count']) + " flow chains verified in target state",
        "0% (pre-migration)",
        "Tower Architect"
    ])
    write_tab(wb, "Success Criteria", TAB_DEFS["Success Criteria"], criteria)

    # 3. NFRs
    profile = tower_ctx.get("nfr_profile", "transactional_highvol")
    nfr_rows = []
    for cat, req, sla, priority, notes in NFR_PROFILES.get(profile, NFR_PROFILES["transactional_highvol"]):
        nfr_rows.append([cat, req, sla, priority, notes])
    write_tab(wb, "NFRs", TAB_DEFS["NFRs"], nfr_rows)

    # 4. Security Controls
    sec_rows = []
    for concern, approach, policy, owner in STANDARD_SECURITY:
        sec_rows.append([concern, approach, policy, owner, ""])
    write_tab(wb, "Security Controls", TAB_DEFS["Security Controls"], sec_rows)

    # 5. SAP Dev Status
    sap_rows = []
    for obj_type, obj_name, env, count, status, notes in STANDARD_SAP_STATUS:
        sap_rows.append([obj_type, obj_name, env, count, status, notes])
    write_tab(wb, "SAP Dev Status", TAB_DEFS["SAP Dev Status"], sap_rows)

    # 6. Recommendations
    rec_rows = generate_recommendations(tower_code, cap_id, cap_name, flow_ctx)
    write_tab(wb, "Recommendations", TAB_DEFS["Recommendations"], rec_rows)

    wb.save(str(xlsx_path))
    return "backfilled-6-tabs (" + str(flow_ctx['hop_count']) + " hops, " + str(len(flow_ctx['systems'])) + " systems)"


# ---------------------------------------------------------------------------
# Discovery + tower.yaml loading
# ---------------------------------------------------------------------------
def discover_capabilities(tower_dir: Path) -> list[Path]:
    """Find all capability folders that have input/data directory."""
    caps = []
    for root, dirs, files in os.walk(str(tower_dir)):
        d = Path(root)
        if (d / "input" / "data").is_dir():
            # Check there are xlsx files in data dir
            if any((d / "input" / "data").glob("*.xlsx")):
                caps.append(d)
    return sorted(caps, key=lambda p: p.name)


def load_cap_names(tower_dir: Path) -> dict:
    """Load capability names from tower.yaml."""
    yml = tower_dir / "tower.yaml"
    if not yml.exists():
        return {}
    try:
        with open(yml, encoding="utf-8") as f:
            data = yaml.safe_load(f)
        result = {}
        for c in data.get("capabilities", []):
            result[c["id"]] = c.get("name", c["id"])
        return result
    except Exception:
        return {}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Backfill supplementary tabs with grounded sample data")
    parser.add_argument("--tower", help="Single tower shortcode")
    parser.add_argument("--cap", help="Single capability ID")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if args.tower:
        tower_dirs = [TOWERS_DIR / args.tower]
        if not tower_dirs[0].is_dir():
            for d in TOWERS_DIR.iterdir():
                if d.is_dir() and d.name.upper() == args.tower.upper():
                    tower_dirs = [d]
                    break
            else:
                print("ERROR: Tower not found: " + args.tower)
                sys.exit(1)
    else:
        tower_dirs = sorted([d for d in TOWERS_DIR.iterdir() if d.is_dir()])

    total_caps = 0
    total_backfilled = 0

    for tower_dir in tower_dirs:
        tower_code = tower_dir.name
        cap_names = load_cap_names(tower_dir)

        print("\n" + "=" * 60)
        print("TOWER: " + tower_code)
        print("=" * 60)

        caps = discover_capabilities(tower_dir)
        if args.cap:
            caps = [c for c in caps if c.name == args.cap]

        for cap_dir in caps:
            total_caps += 1
            cap_id = cap_dir.name
            cap_name = cap_names.get(cap_id, cap_id)
            # Sanitize unicode for console output
            safe_name = cap_name.encode("ascii", errors="replace").decode("ascii")
            data_dir = cap_dir / "input" / "data"

            print("\n  " + cap_id + ": " + safe_name)

            for xlsx_file in sorted(data_dir.glob("*.xlsx")):
                status = backfill_workbook(xlsx_file, tower_code, cap_id, cap_name, args.dry_run)
                print("    " + xlsx_file.name + ": " + status)
                if "backfilled" in status or "would" in status:
                    total_backfilled += 1

    mode = "DRY RUN" if args.dry_run else "COMPLETE"
    print("\n" + "=" * 60)
    print("BACKFILL " + mode)
    print("  Capabilities: " + str(total_caps))
    print("  Workbooks backfilled: " + str(total_backfilled))
    print("=" * 60)


if __name__ == "__main__":
    main()
