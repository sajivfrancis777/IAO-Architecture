"""Backfill extended D/A/T columns (26-47) into existing flow XLSX files.

Reads existing CurrentFlows.xlsx / FutureFlows.xlsx for a capability,
maps each hop's source/target system to realistic Data Architecture,
Technology Architecture, and Endpoint-level attributes, then writes
an updated XLSX with all 47 columns populated.

Usage:
    python scripts/backfill_dat_columns.py --tower FPR --cap DS-020
    python scripts/backfill_dat_columns.py --tower FPR --cap DS-020 --dry-run
"""

from __future__ import annotations

import argparse
import shutil
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

WORKSPACE = Path(__file__).resolve().parent.parent

# ── System-to-platform knowledge base ────────────────────────────
# Maps system names (or substrings) to their technology attributes.
# These are grounded in the actual Intel IDM 2.0 landscape.

SYSTEM_PROFILES: dict[str, dict] = {
    # SAP Systems
    "S/4": {
        "db_platform": "SAP HANA",
        "tech_platform": "SAP S/4HANA On-Premise",
        "default_schema": "SAPHANADB",
        "default_protocol": "RFC / OData",
        "default_auth": "SAP SSO / X.509",
    },
    "S4": {
        "db_platform": "SAP HANA",
        "tech_platform": "SAP S/4HANA On-Premise",
        "default_schema": "SAPHANADB",
        "default_protocol": "RFC / OData",
        "default_auth": "SAP SSO / X.509",
    },
    "SAP BW": {
        "db_platform": "SAP HANA",
        "tech_platform": "SAP BW/4HANA",
        "default_schema": "SAPBWDB",
        "default_protocol": "RFC / ODP",
        "default_auth": "SAP SSO",
    },
    "BW4": {
        "db_platform": "SAP HANA",
        "tech_platform": "SAP BW/4HANA",
        "default_schema": "SAPBWDB",
        "default_protocol": "RFC / ODP",
        "default_auth": "SAP SSO",
    },
    "BPC": {
        "db_platform": "SAP HANA",
        "tech_platform": "SAP BPC (Embedded)",
        "default_schema": "SAPBPC",
        "default_protocol": "RFC / BAPI",
        "default_auth": "SAP SSO",
    },
    "ECC": {
        "db_platform": "Oracle DB",
        "tech_platform": "SAP ECC 6.0 On-Premise",
        "default_schema": "SAPSR3",
        "default_protocol": "RFC / IDoc",
        "default_auth": "SAP SSO / SNC",
    },
    "R/3": {
        "db_platform": "Oracle DB",
        "tech_platform": "SAP R/3 4.7",
        "default_schema": "SAPSR3",
        "default_protocol": "RFC / IDoc",
        "default_auth": "SAP SSO / SNC",
    },
    "SAP CPI": {
        "db_platform": "N/A (iPaaS)",
        "tech_platform": "SAP BTP Cloud",
        "default_schema": "N/A",
        "default_protocol": "REST / SOAP / OData",
        "default_auth": "OAuth 2.0 / X.509",
    },
    "CPI": {
        "db_platform": "N/A (iPaaS)",
        "tech_platform": "SAP BTP Cloud",
        "default_schema": "N/A",
        "default_protocol": "REST / SOAP",
        "default_auth": "OAuth 2.0",
    },
    "SAP BTP": {
        "db_platform": "SAP HANA Cloud",
        "tech_platform": "SAP BTP Cloud",
        "default_schema": "HANACLOUD",
        "default_protocol": "REST / OData",
        "default_auth": "OAuth 2.0",
    },
    "Signavio": {
        "db_platform": "N/A (SaaS)",
        "tech_platform": "SAP Signavio SaaS",
        "default_schema": "N/A",
        "default_protocol": "REST API",
        "default_auth": "OAuth 2.0 / SAML",
    },
    "Fiori": {
        "db_platform": "SAP HANA",
        "tech_platform": "SAP Fiori / Gateway",
        "default_schema": "SAPHANADB",
        "default_protocol": "OData V2/V4",
        "default_auth": "SAP SSO / SAML",
    },
    # Microsoft / Azure
    "Azure": {
        "db_platform": "Azure SQL",
        "tech_platform": "Microsoft Azure Cloud",
        "default_schema": "dbo",
        "default_protocol": "REST / HTTPS",
        "default_auth": "Azure AD / OAuth 2.0",
    },
    "Power BI": {
        "db_platform": "Azure Analysis Services",
        "tech_platform": "Microsoft Power BI SaaS",
        "default_schema": "N/A",
        "default_protocol": "DAX / REST",
        "default_auth": "Azure AD SSO",
    },
    "SharePoint": {
        "db_platform": "Azure SQL (managed)",
        "tech_platform": "Microsoft 365 SaaS",
        "default_schema": "N/A",
        "default_protocol": "Microsoft Graph API",
        "default_auth": "Azure AD / OAuth 2.0",
    },
    "Teams": {
        "db_platform": "N/A (SaaS)",
        "tech_platform": "Microsoft 365 SaaS",
        "default_schema": "N/A",
        "default_protocol": "Microsoft Graph API",
        "default_auth": "Azure AD SSO",
    },
    # Intel Internal
    "ICOST": {
        "db_platform": "Oracle DB",
        "tech_platform": "Intel Custom On-Premise",
        "default_schema": "ICOST_PROD",
        "default_protocol": "JDBC / File",
        "default_auth": "NTLM / Kerberos",
    },
    "MES": {
        "db_platform": "Oracle DB",
        "tech_platform": "Intel Manufacturing On-Premise",
        "default_schema": "MES_PROD",
        "default_protocol": "REST / MQTT",
        "default_auth": "Kerberos / X.509",
    },
    "PDH": {
        "db_platform": "Oracle DB",
        "tech_platform": "Intel Manufacturing On-Premise",
        "default_schema": "PDH_PROD",
        "default_protocol": "JDBC / File",
        "default_auth": "Kerberos",
    },
    "XEUS": {
        "db_platform": "Oracle DB",
        "tech_platform": "Intel Middleware On-Premise",
        "default_schema": "XEUS_PROD",
        "default_protocol": "JMS / File",
        "default_auth": "Kerberos / NTLM",
    },
    "Hyperion": {
        "db_platform": "Oracle Essbase",
        "tech_platform": "Oracle Hyperion On-Premise",
        "default_schema": "HPR_PROD",
        "default_protocol": "Essbase API / File",
        "default_auth": "NTLM / LDAP",
    },
    "Anaplan": {
        "db_platform": "N/A (SaaS)",
        "tech_platform": "Anaplan Cloud SaaS",
        "default_schema": "N/A",
        "default_protocol": "REST API",
        "default_auth": "OAuth 2.0 / SAML SSO",
    },
    "Coupa": {
        "db_platform": "N/A (SaaS)",
        "tech_platform": "Coupa Cloud SaaS",
        "default_schema": "N/A",
        "default_protocol": "REST / cXML",
        "default_auth": "OAuth 2.0 / API Key",
    },
    "Ariba": {
        "db_platform": "SAP HANA Cloud",
        "tech_platform": "SAP Ariba Cloud SaaS",
        "default_schema": "N/A",
        "default_protocol": "REST / cXML",
        "default_auth": "OAuth 2.0 / SAML",
    },
    "Concur": {
        "db_platform": "N/A (SaaS)",
        "tech_platform": "SAP Concur Cloud SaaS",
        "default_schema": "N/A",
        "default_protocol": "REST API",
        "default_auth": "OAuth 2.0 / SAML SSO",
    },
    "Smartsheet": {
        "db_platform": "N/A (SaaS)",
        "tech_platform": "Smartsheet Cloud SaaS",
        "default_schema": "N/A",
        "default_protocol": "REST API",
        "default_auth": "Bearer Token / OAuth 2.0",
    },
    "JIRA": {
        "db_platform": "PostgreSQL",
        "tech_platform": "Atlassian Server On-Premise",
        "default_schema": "jiradb",
        "default_protocol": "REST API",
        "default_auth": "PAT / Bearer Token",
    },
    "ServiceNow": {
        "db_platform": "MariaDB (managed)",
        "tech_platform": "ServiceNow Cloud SaaS",
        "default_schema": "N/A",
        "default_protocol": "REST / SOAP",
        "default_auth": "OAuth 2.0 / SAML SSO",
    },
    "Salesforce": {
        "db_platform": "N/A (SaaS)",
        "tech_platform": "Salesforce Cloud SaaS",
        "default_schema": "N/A",
        "default_protocol": "REST / SOAP",
        "default_auth": "OAuth 2.0 / SAML",
    },
    # Intel-specific systems (IDM 2.0 landscape)
    "CFIN": {
        "db_platform": "SAP HANA",
        "tech_platform": "SAP S/4HANA Central Finance",
        "default_schema": "SAPHANADB",
        "default_protocol": "RFC / AIF",
        "default_auth": "SAP SSO / X.509",
    },
    "Corp / IP S/4": {
        "db_platform": "SAP HANA",
        "tech_platform": "SAP S/4HANA On-Premise",
        "default_schema": "SAPHANADB",
        "default_protocol": "RFC / OData",
        "default_auth": "SAP SSO / X.509",
    },
    "IF S/4": {
        "db_platform": "SAP HANA",
        "tech_platform": "SAP S/4HANA On-Premise",
        "default_schema": "SAPHANADB",
        "default_protocol": "RFC / OData",
        "default_auth": "SAP SSO / X.509",
    },
    "SAP S/4 MDG": {
        "db_platform": "SAP HANA",
        "tech_platform": "SAP MDG On-Premise",
        "default_schema": "SAPHANADB",
        "default_protocol": "RFC / OData",
        "default_auth": "SAP SSO / X.509",
    },
    "SAP IBP": {
        "db_platform": "SAP HANA Cloud",
        "tech_platform": "SAP IBP Cloud SaaS",
        "default_schema": "IBP_HANA",
        "default_protocol": "REST / OData / CPI",
        "default_auth": "OAuth 2.0 / SAML",
    },
    "SAP PAPM": {
        "db_platform": "SAP HANA",
        "tech_platform": "SAP PaPM On-Premise",
        "default_schema": "SAPHANADB",
        "default_protocol": "RFC / HANA Views",
        "default_auth": "SAP SSO",
    },
    "SAP SAC": {
        "db_platform": "SAP HANA Cloud",
        "tech_platform": "SAP Analytics Cloud SaaS",
        "default_schema": "N/A",
        "default_protocol": "REST / Import API",
        "default_auth": "OAuth 2.0 / SAML SSO",
    },
    "SAP PO": {
        "db_platform": "Oracle DB",
        "tech_platform": "SAP Process Orchestration On-Premise",
        "default_schema": "SAPPO",
        "default_protocol": "RFC / SOAP / IDoc",
        "default_auth": "X.509 / Basic Auth",
    },
    "SAP BOBJ": {
        "db_platform": "SAP HANA",
        "tech_platform": "SAP BusinessObjects On-Premise",
        "default_schema": "BOBJ_CMS",
        "default_protocol": "JDBC / BICS",
        "default_auth": "SAP SSO / Enterprise",
    },
    "BOBJ": {
        "db_platform": "SAP HANA",
        "tech_platform": "SAP BusinessObjects On-Premise",
        "default_schema": "BOBJ_CMS",
        "default_protocol": "JDBC / BICS",
        "default_auth": "SAP SSO / Enterprise",
    },
    "SAP BODS": {
        "db_platform": "Oracle DB",
        "tech_platform": "SAP Data Services On-Premise",
        "default_schema": "BODS_REPO",
        "default_protocol": "JDBC / File / RFC",
        "default_auth": "OS Auth / DB Auth",
    },
    "Finance HANA": {
        "db_platform": "SAP HANA",
        "tech_platform": "SAP HANA Sidecar On-Premise",
        "default_schema": "FINANCE_DB",
        "default_protocol": "HANA SQL / SDA",
        "default_auth": "SAP SSO / X.509",
    },
    "Legacy MDG": {
        "db_platform": "Oracle DB",
        "tech_platform": "SAP ECC MDG On-Premise",
        "default_schema": "SAPSR3",
        "default_protocol": "RFC / IDoc",
        "default_auth": "SAP SSO",
    },
    "MES 300": {
        "db_platform": "Oracle DB",
        "tech_platform": "Intel MES On-Premise",
        "default_schema": "MES300_PROD",
        "default_protocol": "REST / File",
        "default_auth": "Kerberos / X.509",
    },
    "WorkStream": {
        "db_platform": "Oracle DB",
        "tech_platform": "Intel MES On-Premise",
        "default_schema": "WS_PROD",
        "default_protocol": "REST / File",
        "default_auth": "Kerberos",
    },
    "MARS": {
        "db_platform": "SQL Server",
        "tech_platform": "Intel Custom On-Premise",
        "default_schema": "MARS_PROD",
        "default_protocol": "REST / File",
        "default_auth": "Kerberos / NTLM",
    },
    "APIGEE": {
        "db_platform": "N/A (API Gateway)",
        "tech_platform": "Google Apigee Cloud",
        "default_schema": "N/A",
        "default_protocol": "REST / GraphQL",
        "default_auth": "OAuth 2.0 / API Key",
    },
    "Azure ADF": {
        "db_platform": "N/A (ETL)",
        "tech_platform": "Azure Data Factory Cloud",
        "default_schema": "N/A",
        "default_protocol": "REST / JDBC / File",
        "default_auth": "Azure AD / Managed Identity",
    },
    "ECA": {
        "db_platform": "Azure Data Lake (ADLS)",
        "tech_platform": "Intel ECA Platform Azure Cloud",
        "default_schema": "ECA_LAKE",
        "default_protocol": "REST / Spark SQL",
        "default_auth": "Azure AD / Service Principal",
    },
    "ECA-ADLS": {
        "db_platform": "Azure Data Lake (ADLS)",
        "tech_platform": "Intel ECA Platform Azure Cloud",
        "default_schema": "ECA_RAW",
        "default_protocol": "ABFS / REST",
        "default_auth": "Azure AD / Managed Identity",
    },
    "ECA-DataBricks": {
        "db_platform": "Delta Lake",
        "tech_platform": "Databricks on ECA Azure Cloud",
        "default_schema": "unity_catalog.eca",
        "default_protocol": "Spark SQL / REST",
        "default_auth": "Azure AD / PAT",
    },
    "ECA-SnowFlake": {
        "db_platform": "Snowflake Cloud DW",
        "tech_platform": "Snowflake on ECA Cloud",
        "default_schema": "ECA_DW",
        "default_protocol": "SQL / REST",
        "default_auth": "Azure AD / Key Pair",
    },
    "EDW": {
        "db_platform": "Teradata / Oracle DB",
        "tech_platform": "Intel EDW On-Premise",
        "default_schema": "EDW_PROD",
        "default_protocol": "JDBC / File",
        "default_auth": "LDAP / Kerberos",
    },
    "FCA": {
        "db_platform": "SQL Server",
        "tech_platform": "Intel Custom On-Premise",
        "default_schema": "FCA_PROD",
        "default_protocol": "JDBC / File",
        "default_auth": "Kerberos / NTLM",
    },
    "FCS": {
        "db_platform": "SQL Server",
        "tech_platform": "Intel Custom On-Premise",
        "default_schema": "FCS_PROD",
        "default_protocol": "JDBC / File",
        "default_auth": "Kerberos / NTLM",
    },
    "PDF-SMH": {
        "db_platform": "N/A (Middleware)",
        "tech_platform": "Intel PDF-SMH Middleware On-Premise",
        "default_schema": "N/A",
        "default_protocol": "File / JMS",
        "default_auth": "Service Account / Kerberos",
    },
    "SideCar": {
        "db_platform": "SAP HANA",
        "tech_platform": "SAP HANA Sidecar On-Premise",
        "default_schema": "SIDECAR_DB",
        "default_protocol": "HANA SQL / SDA",
        "default_auth": "SAP SSO",
    },
    "PDH": {
        "db_platform": "Oracle DB",
        "tech_platform": "Intel PDH On-Premise",
        "default_schema": "PDH_PROD",
        "default_protocol": "REST / File",
        "default_auth": "Kerberos",
    },
    "IF PDH": {
        "db_platform": "Oracle DB",
        "tech_platform": "Intel PDH On-Premise",
        "default_schema": "PDH_IF",
        "default_protocol": "REST / File",
        "default_auth": "Kerberos",
    },
    "IP PDH": {
        "db_platform": "Oracle DB",
        "tech_platform": "Intel PDH On-Premise",
        "default_schema": "PDH_IP",
        "default_protocol": "REST / File",
        "default_auth": "Kerberos",
    },
    "PDM Translator": {
        "db_platform": "SQL Server",
        "tech_platform": "Intel PDM On-Premise",
        "default_schema": "PDM_TRANS",
        "default_protocol": "File / REST",
        "default_auth": "Kerberos",
    },
    "Capacity Forecast": {
        "db_platform": "SQL Server",
        "tech_platform": "Intel Custom On-Premise",
        "default_schema": "CAP_FCST",
        "default_protocol": "JDBC / File",
        "default_auth": "Kerberos / NTLM",
    },
    "GraphiteConnect": {
        "db_platform": "N/A (SaaS)",
        "tech_platform": "GraphiteConnect Cloud SaaS",
        "default_schema": "N/A",
        "default_protocol": "REST / EDI",
        "default_auth": "OAuth 2.0 / API Key",
    },
    "ECM": {
        "db_platform": "Oracle DB",
        "tech_platform": "PTC Windchill On-Premise",
        "default_schema": "WINDCHILL_PROD",
        "default_protocol": "REST / SOAP",
        "default_auth": "LDAP / SSO",
    },
    "Windchill": {
        "db_platform": "Oracle DB",
        "tech_platform": "PTC Windchill On-Premise",
        "default_schema": "WINDCHILL_PROD",
        "default_protocol": "REST / SOAP",
        "default_auth": "LDAP / SSO",
    },
    "Blue Yonder": {
        "db_platform": "SQL Server",
        "tech_platform": "Blue Yonder Cloud SaaS",
        "default_schema": "BY_PROD",
        "default_protocol": "REST / File",
        "default_auth": "OAuth 2.0 / SAML",
    },
    "ICS": {
        "db_platform": "Oracle DB",
        "tech_platform": "Intel ICS On-Premise",
        "default_schema": "ICS_PROD",
        "default_protocol": "REST / File",
        "default_auth": "Kerberos",
    },
    "Phoenix": {
        "db_platform": "Oracle DB",
        "tech_platform": "Intel ICS (Phoenix) On-Premise",
        "default_schema": "PHOENIX_PROD",
        "default_protocol": "REST / File",
        "default_auth": "Kerberos",
    },
    "PEGA": {
        "db_platform": "PostgreSQL",
        "tech_platform": "Pega Cloud SaaS",
        "default_schema": "PEGA_PROD",
        "default_protocol": "REST / SOAP",
        "default_auth": "OAuth 2.0 / SAML",
    },
    "SPEED": {
        "db_platform": "SQL Server",
        "tech_platform": "Intel Custom On-Premise",
        "default_schema": "SPEED_PROD",
        "default_protocol": "REST / File",
        "default_auth": "Kerberos / NTLM",
    },
    "EATS": {
        "db_platform": "SQL Server",
        "tech_platform": "Intel Custom On-Premise",
        "default_schema": "EATS_PROD",
        "default_protocol": "REST / File",
        "default_auth": "Kerberos / NTLM",
    },
    "WSPW": {
        "db_platform": "SQL Server",
        "tech_platform": "Intel Custom On-Premise",
        "default_schema": "WSPW_PROD",
        "default_protocol": "JDBC / File",
        "default_auth": "Kerberos",
    },
    "SCS": {
        "db_platform": "SQL Server",
        "tech_platform": "Intel Custom On-Premise",
        "default_schema": "SCS_PROD",
        "default_protocol": "REST / File",
        "default_auth": "Kerberos / NTLM",
    },
    "ATCR": {
        "db_platform": "SQL Server",
        "tech_platform": "Intel Custom On-Premise",
        "default_schema": "ATCR_PROD",
        "default_protocol": "REST / File",
        "default_auth": "Kerberos / NTLM",
    },
    "CIBR": {
        "db_platform": "SQL Server",
        "tech_platform": "Intel Custom On-Premise",
        "default_schema": "CIBR_PROD",
        "default_protocol": "JDBC / File",
        "default_auth": "Kerberos",
    },
    "DMOCR": {
        "db_platform": "SQL Server",
        "tech_platform": "Intel Custom On-Premise",
        "default_schema": "DMOCR_PROD",
        "default_protocol": "JDBC / File",
        "default_auth": "Kerberos",
    },
    "DXCR": {
        "db_platform": "SQL Server",
        "tech_platform": "Intel Custom On-Premise",
        "default_schema": "DXCR_PROD",
        "default_protocol": "JDBC / File",
        "default_auth": "Kerberos",
    },
    "Power BI": {
        "db_platform": "Azure Analysis Services",
        "tech_platform": "Microsoft Power BI SaaS",
        "default_schema": "N/A",
        "default_protocol": "DAX / REST",
        "default_auth": "Azure AD SSO",
    },
    "DARC": {
        "db_platform": "Azure Analysis Services",
        "tech_platform": "Microsoft Power BI (DARC) SaaS",
        "default_schema": "DARC_MODEL",
        "default_protocol": "DAX / REST",
        "default_auth": "Azure AD SSO",
    },
    "Snowflake": {
        "db_platform": "Snowflake Cloud DW",
        "tech_platform": "Snowflake Cloud SaaS",
        "default_schema": "INTEL_DW",
        "default_protocol": "SQL / REST",
        "default_auth": "Azure AD / Key Pair",
    },
    "Databricks": {
        "db_platform": "Delta Lake",
        "tech_platform": "Databricks Cloud SaaS",
        "default_schema": "unity_catalog",
        "default_protocol": "SQL / REST / Spark",
        "default_auth": "Azure AD / PAT",
    },
    "Kafka": {
        "db_platform": "N/A (Event Broker)",
        "tech_platform": "Apache Kafka On-Premise",
        "default_schema": "N/A",
        "default_protocol": "Kafka Protocol / AMQP",
        "default_auth": "mTLS / SASL",
    },
    "MuleSoft": {
        "db_platform": "N/A (iPaaS)",
        "tech_platform": "MuleSoft Anypoint Cloud",
        "default_schema": "N/A",
        "default_protocol": "REST / SOAP",
        "default_auth": "OAuth 2.0 / Client Creds",
    },
}

# Data entity inference: interface tech / data description → entity + format
DATA_ENTITY_MAP: dict[str, dict] = {
    "cost": {"entity": "Cost Element", "format": "IDoc / Flat File", "class": "Intel Confidential", "type": "Transaction"},
    "invoice": {"entity": "Invoice", "format": "IDoc / XML", "class": "Intel Confidential", "type": "Transaction"},
    "purchase": {"entity": "Purchase Order", "format": "IDoc / XML", "class": "Intel Confidential", "type": "Transaction"},
    "sales": {"entity": "Sales Order", "format": "IDoc / XML", "class": "Intel Confidential", "type": "Transaction"},
    "material": {"entity": "Material Master", "format": "IDoc / CSV", "class": "Intel Restricted", "type": "Master"},
    "vendor": {"entity": "Vendor Master", "format": "IDoc / CSV", "class": "Intel Confidential", "type": "Master"},
    "customer": {"entity": "Customer Master", "format": "IDoc / CSV", "class": "Intel Confidential", "type": "Master"},
    "gl": {"entity": "GL Account", "format": "BAPI / OData", "class": "Intel Confidential", "type": "Master"},
    "journal": {"entity": "Journal Entry", "format": "BAPI / IDoc", "class": "Intel Confidential", "type": "Transaction"},
    "payment": {"entity": "Payment", "format": "IDoc / XML", "class": "Intel Secret", "type": "Transaction"},
    "delivery": {"entity": "Delivery", "format": "IDoc / XML", "class": "Intel Confidential", "type": "Transaction"},
    "inventory": {"entity": "Inventory", "format": "IDoc / CSV", "class": "Intel Confidential", "type": "Transaction"},
    "bom": {"entity": "Bill of Materials", "format": "IDoc / CSV", "class": "Intel Restricted", "type": "Master"},
    "routing": {"entity": "Routing", "format": "IDoc / CSV", "class": "Intel Restricted", "type": "Master"},
    "plan": {"entity": "Plan Data", "format": "CSV / JSON", "class": "Intel Confidential", "type": "Transaction"},
    "forecast": {"entity": "Forecast Data", "format": "CSV / JSON", "class": "Intel Confidential", "type": "Transaction"},
    "budget": {"entity": "Budget Data", "format": "CSV / BAPI", "class": "Intel Confidential", "type": "Transaction"},
    "exchange": {"entity": "Exchange Rate", "format": "CSV / API", "class": "Public", "type": "Master"},
    "tax": {"entity": "Tax Data", "format": "XML / API", "class": "Intel Confidential", "type": "Master"},
    "asset": {"entity": "Fixed Asset", "format": "IDoc / BAPI", "class": "Intel Confidential", "type": "Master"},
    "employee": {"entity": "Employee Master", "format": "IDoc / CSV", "class": "Intel Secret", "type": "Master"},
    "profit": {"entity": "Profitability Data", "format": "BAPI / Flat File", "class": "Intel Confidential", "type": "Transaction"},
    "consol": {"entity": "Consolidation Data", "format": "CSV / API", "class": "Intel Confidential", "type": "Transaction"},
    "report": {"entity": "Financial Report", "format": "CSV / PDF", "class": "Intel Confidential", "type": "Transaction"},
    "production": {"entity": "Production Order", "format": "IDoc / RFC", "class": "Intel Confidential", "type": "Transaction"},
    "quality": {"entity": "Quality Record", "format": "IDoc / CSV", "class": "Intel Confidential", "type": "Transaction"},
    "shipment": {"entity": "Shipment", "format": "IDoc / EDI", "class": "Intel Confidential", "type": "Transaction"},
    "batch": {"entity": "Batch Master", "format": "IDoc / CSV", "class": "Intel Restricted", "type": "Master"},
    "warehouse": {"entity": "Warehouse Order", "format": "RFC / OData", "class": "Intel Confidential", "type": "Transaction"},
}

INTEGRATION_PATTERN_MAP: dict[str, str] = {
    "API": "Point-to-Point",
    "REST": "Point-to-Point",
    "OData": "Point-to-Point",
    "RFC": "Point-to-Point",
    "BAPI": "Point-to-Point",
    "IDoc": "Pub-Sub",
    "File": "File Transfer",
    "SFTP": "File Transfer",
    "FTP": "File Transfer",
    "Flat File": "File Transfer",
    "CSV": "File Transfer",
    "EDI": "Pub-Sub",
    "Kafka": "Event-Driven",
    "Event": "Event-Driven",
    "MQTT": "Event-Driven",
    "AMQP": "Event-Driven",
    "ETL": "ETL / Batch",
    "Batch": "ETL / Batch",
    "Direct": "Point-to-Point",
    "JDBC": "Point-to-Point",
    "SOAP": "Point-to-Point",
}


def _resolve_profile(system_name: str) -> dict:
    """Find the best matching system profile.
    
    Matching priority: exact > longest substring > fallback.
    """
    if not system_name:
        return {}
    name_upper = system_name.upper()
    # Exact match first
    for key, profile in SYSTEM_PROFILES.items():
        if key.upper() == name_upper:
            return profile
    # Longest substring match (prefer more specific keys)
    best_key, best_profile, best_len = None, None, 0
    for key, profile in SYSTEM_PROFILES.items():
        if key.upper() in name_upper and len(key) > best_len:
            best_key, best_profile, best_len = key, profile, len(key)
    if best_profile:
        return best_profile
    # Default
    return {
        "db_platform": "Unknown DB",
        "tech_platform": f"{system_name} Platform",
        "default_schema": "N/A",
        "default_protocol": "HTTPS",
        "default_auth": "SSO",
    }


def _infer_data_entity(data_desc: str, flow_purpose: str, interface: str) -> dict:
    """Infer data entity from data description / flow purpose."""
    search_text = f"{data_desc} {flow_purpose} {interface}".lower()
    for keyword, info in DATA_ENTITY_MAP.items():
        if keyword in search_text:
            return info
    return {"entity": "Business Data", "format": "JSON / CSV", "class": "Intel Confidential", "type": "Transaction"}


def _infer_integration_pattern(interface_tech: str) -> str:
    """Infer integration pattern from interface technology."""
    if not interface_tech:
        return "Point-to-Point"
    for keyword, pattern in INTEGRATION_PATTERN_MAP.items():
        if keyword.lower() in interface_tech.lower():
            return pattern
    return "Point-to-Point"


def _infer_volume(data_type: str, frequency: str) -> str:
    """Estimate data volume from type and frequency."""
    freq_lower = (frequency or "").lower()
    if "real" in freq_lower or "event" in freq_lower:
        return "~10K events/hour"
    if "daily" in freq_lower or "batch" in freq_lower:
        return "~50K-500K records/day"
    if "weekly" in freq_lower:
        return "~100K records/week"
    if "monthly" in freq_lower:
        return "~1M records/month"
    if data_type == "Master":
        return "~10K-100K records (full refresh)"
    return "~10K-100K records/day"


def _infer_middleware(interface_tech: str, src_profile: dict, tgt_profile: dict) -> str:
    """Infer middleware from interface tech and platforms."""
    iface = (interface_tech or "").lower()
    if "cpi" in iface or "cloud" in iface:
        return "SAP CPI"
    if "mule" in iface:
        return "MuleSoft Anypoint"
    if "kafka" in iface or "event" in iface:
        return "Apache Kafka"
    if "xeus" in iface:
        return "XEUS"
    # If both SAP, likely CPI or direct
    src_sap = "SAP" in src_profile.get("tech_platform", "")
    tgt_sap = "SAP" in tgt_profile.get("tech_platform", "")
    if src_sap and tgt_sap:
        return "Direct / SAP CPI"
    if src_sap or tgt_sap:
        return "SAP CPI"
    return "Direct"


# ── Main backfill logic ──────────────────────────────────────────

EXTENDED_HEADERS = [
    # Cols 26-31: Data Architecture
    "Data Entity", "Data Format", "Data Classification",
    "Data Volume", "Master/Transaction", "Data Lineage Notes",
    # Cols 32-37: Technology Architecture
    "Integration Pattern", "Middleware / Platform", "Protocol",
    "Auth Method", "Environment Scope", "SLA / Latency",
    # Cols 38-41: Interface Architecture
    "Interface ID", "Interface Type", "Error Handling", "Monitoring",
    # Cols 42-47: Endpoint-level
    "Source DB Platform", "Target DB Platform",
    "Source Schema/Object", "Target Schema/Object",
    "Source Tech Platform", "Target Tech Platform",
]

BASE_COL_COUNT = 25


def backfill_xlsx(xlsx_path: Path, dry_run: bool = False) -> int:
    """Add/populate extended columns (26-47) in a flow XLSX file.
    
    Returns number of rows updated.
    """
    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active

    # Determine existing column count
    header_row = [cell.value for cell in ws[1]]
    existing_cols = len([h for h in header_row if h])

    # Add extended headers if missing
    if existing_cols <= BASE_COL_COUNT:
        for i, hdr in enumerate(EXTENDED_HEADERS):
            col_idx = BASE_COL_COUNT + 1 + i
            ws.cell(row=1, column=col_idx, value=hdr)
    
    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        # Read base columns
        flow_chain = ws.cell(row=row_idx, column=1).value
        if not flow_chain:
            continue

        source_sys = str(ws.cell(row=row_idx, column=3).value or "")
        target_sys = str(ws.cell(row=row_idx, column=5).value or "")
        interface = str(ws.cell(row=row_idx, column=7).value or "")
        frequency = str(ws.cell(row=row_idx, column=9).value or "")
        data_desc = str(ws.cell(row=row_idx, column=10).value or "")
        flow_purpose = str(ws.cell(row=row_idx, column=11).value or "")
        hop_num = int(ws.cell(row=row_idx, column=2).value or 1)

        # Skip if already has extended data (check col 26 - Data Entity)
        existing_entity = ws.cell(row=row_idx, column=26).value
        if existing_entity and not str(existing_entity).startswith("e.g."):
            continue

        # Resolve profiles
        src_prof = _resolve_profile(source_sys)
        tgt_prof = _resolve_profile(target_sys)
        data_info = _infer_data_entity(data_desc, flow_purpose, interface)
        int_pattern = _infer_integration_pattern(interface)
        middleware = _infer_middleware(interface, src_prof, tgt_prof)
        volume = _infer_volume(data_info["type"], frequency)

        # Build schema/object names from data description
        src_schema = src_prof.get("default_schema", "N/A")
        tgt_schema = tgt_prof.get("default_schema", "N/A")
        if data_desc:
            desc_slug = data_desc.replace(" ", "_").replace("/", "_")[:30]
            if src_schema != "N/A":
                src_schema = f"{src_schema}.{desc_slug}"
            if tgt_schema != "N/A":
                tgt_schema = f"{tgt_schema}.{desc_slug}"

        # Generate interface ID
        chain_slug = str(flow_chain).replace(" ", "")[:10]
        iface_id = f"IF-{chain_slug}-{hop_num:02d}" if hop_num else f"IF-{chain_slug}"

        # Determine SLA
        if "real" in frequency.lower():
            sla = "< 1 second (real-time)"
        elif "daily" in frequency.lower() or "batch" in frequency.lower():
            sla = "< 4 hours (batch window)"
        elif "weekly" in frequency.lower() or "monthly" in frequency.lower():
            sla = "< 24 hours"
        else:
            sla = "< 1 hour (near real-time)"

        # Write extended columns (26-47)
        values = [
            # Data Architecture (26-31)
            data_info["entity"],                          # 26: Data Entity
            data_info["format"],                          # 27: Data Format
            data_info["class"],                           # 28: Data Classification
            volume,                                       # 29: Data Volume
            data_info["type"],                            # 30: Master/Transaction
            f"{source_sys} → {target_sys}: {data_desc}",  # 31: Data Lineage Notes
            # Technology Architecture (32-37)
            int_pattern,                                  # 32: Integration Pattern
            middleware,                                   # 33: Middleware / Platform
            src_prof.get("default_protocol", "HTTPS"),    # 34: Protocol
            src_prof.get("default_auth", "SSO"),          # 35: Auth Method
            "DEV, QAS, PRD",                              # 36: Environment Scope
            sla,                                          # 37: SLA / Latency
            # Interface Architecture (38-41)
            iface_id,                                     # 38: Interface ID
            int_pattern,                                  # 39: Interface Type
            "Retry 3x → Dead Letter Queue → Alert",      # 40: Error Handling
            "SAP CCMS / Azure Monitor / Splunk",          # 41: Monitoring
            # Endpoint-level (42-47)
            src_prof.get("db_platform", "Unknown"),       # 42: Source DB Platform
            tgt_prof.get("db_platform", "Unknown"),       # 43: Target DB Platform
            src_schema,                                   # 44: Source Schema/Object
            tgt_schema,                                   # 45: Target Schema/Object
            src_prof.get("tech_platform", "Unknown"),     # 46: Source Tech Platform
            tgt_prof.get("tech_platform", "Unknown"),     # 47: Target Tech Platform
        ]

        for i, val in enumerate(values):
            ws.cell(row=row_idx, column=BASE_COL_COUNT + 1 + i, value=val)
        updated += 1

    if not dry_run and updated > 0:
        # Backup original
        backup = xlsx_path.with_suffix(".xlsx.bak")
        if not backup.exists():
            shutil.copy2(xlsx_path, backup)
        wb.save(str(xlsx_path))
        print(f"  Updated {updated} rows in {xlsx_path.name} (backup: {backup.name})")
    else:
        print(f"  [DRY-RUN] Would update {updated} rows in {xlsx_path.name}")

    return updated


# ── CLI ──────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Backfill D/A/T columns (26-47) in flow XLSX files")
    parser.add_argument("--tower", required=True, help="Tower shortcode (FPR, E2E, etc.)")
    parser.add_argument("--cap", default="", help="Capability ID (DS-020). Omit for all caps in tower.")
    parser.add_argument("--dry-run", action="store_true", help="Preview changes without writing")
    args = parser.parse_args()

    towers_dir = WORKSPACE / "towers"
    tower_dir = towers_dir / args.tower
    if not tower_dir.exists():
        # Search case-insensitively
        for d in towers_dir.iterdir():
            if d.is_dir() and d.name.upper() == args.tower.upper():
                tower_dir = d
                break

    if not tower_dir.exists():
        print(f"Tower directory not found: {tower_dir}")
        return

    # Find capability dirs
    if args.cap:
        cap_dirs = list(tower_dir.rglob(args.cap))
        cap_dirs = [d for d in cap_dirs if d.is_dir()]
    else:
        import re
        cap_dirs = [d for d in tower_dir.rglob("*") if d.is_dir() and re.match(r"^[A-Z]+-\d+$", d.name, re.IGNORECASE)]

    if not cap_dirs:
        print(f"No capability directories found for {args.tower}/{args.cap}")
        return

    total = 0
    for cap_dir in sorted(cap_dirs):
        data_dir = cap_dir / "input" / "data"
        if not data_dir.exists():
            continue

        print(f"\n{cap_dir.name}:")
        for flow_file in sorted(data_dir.glob("*Flows.xlsx")):
            total += backfill_xlsx(flow_file, dry_run=args.dry_run)

    print(f"\nTotal rows updated: {total}")


if __name__ == "__main__":
    main()
