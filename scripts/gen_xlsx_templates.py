"""gen_xlsx_templates.py — Generate multi-tab Excel input templates for capabilities.

Creates the blank .xlsx templates with all 8 tabs, proper headers, column widths,
header formatting, and helper text rows. These serve as the input workbooks that
tower architects fill in.

Usage:
    python gen_xlsx_templates.py                           # generate template pair in templates/
    python gen_xlsx_templates.py --deploy --tower FPR      # deploy to all FPR capabilities
    python gen_xlsx_templates.py --deploy                  # deploy to ALL capabilities
    python gen_xlsx_templates.py --deploy --dry-run        # preview deployment
    python gen_xlsx_templates.py --migrate --tower FPR     # migrate existing CSVs into xlsx
    python gen_xlsx_templates.py --migrate                 # migrate ALL capabilities
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
WORKSPACE = Path(__file__).resolve().parent.parent
TOWERS_DIR = WORKSPACE / "towers"
TEMPLATES_DIR = WORKSPACE / "templates"

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(
    bottom=Side(style="thin", color="000000"),
    right=Side(style="thin", color="D9D9D9"),
)

HINT_FONT = Font(name="Calibri", italic=True, size=10, color="808080")
HINT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Group header fills for the Flows tab column groups
GROUP_FILLS = {
    "base": PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),       # dark blue
    "data": PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"),        # medium blue
    "tech": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),        # lighter blue
    "iface": PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid"),       # even lighter
    "endpoint": PatternFill(start_color="9DC3E6", end_color="9DC3E6", fill_type="solid"),    # lightest
}


# ---------------------------------------------------------------------------
# Tab definitions: (header_columns, hint_row_values, column_widths)
# ---------------------------------------------------------------------------

# Tab 1: Flows — the 47-column integration flow data
FLOWS_COLUMNS = [
    # Base 25 (cols 1-25) — "base" group
    ("Flow Chain", 18, "base"),
    ("Hop #", 7, "base"),
    ("Source System", 18, "base"),
    ("Source Lane", 18, "base"),
    ("Target System", 18, "base"),
    ("Target Lane", 18, "base"),
    ("Interface / Technology", 22, "base"),
    ("Direction", 10, "base"),
    ("Frequency", 14, "base"),
    ("Data Description", 25, "base"),
    ("Flow Purpose", 30, "base"),
    ("Notes / Corrections", 25, "base"),
    ("Process/System Owner", 20, "base"),
    ("Data Owner", 18, "base"),
    ("Applicable Scope", 18, "base"),
    ("Src Web Address", 22, "base"),
    ("Src Business Owner", 18, "base"),
    ("Src Product Owner", 18, "base"),
    ("Src Product Owner Email", 22, "base"),
    ("Src IAPM URL", 30, "base"),
    ("Tgt Web Address", 22, "base"),
    ("Tgt Business Owner", 18, "base"),
    ("Tgt Product Owner", 18, "base"),
    ("Tgt Product Owner Email", 22, "base"),
    ("Tgt IAPM URL", 30, "base"),
    # Data Architecture (cols 26-31) — "data" group
    ("Data Entity", 18, "data"),
    ("Data Format", 14, "data"),
    ("Data Classification", 18, "data"),
    ("Data Volume", 14, "data"),
    ("Master/Transaction", 18, "data"),
    ("Data Lineage Notes", 25, "data"),
    # Technology Architecture (cols 32-37) — "tech" group
    ("Integration Pattern", 18, "tech"),
    ("Middleware / Platform", 20, "tech"),
    ("Protocol", 14, "tech"),
    ("Auth Method", 14, "tech"),
    ("Environment Scope", 16, "tech"),
    ("SLA / Latency", 14, "tech"),
    # Interface Architecture (cols 38-41) — "iface" group
    ("Interface ID", 14, "iface"),
    ("Interface Type", 14, "iface"),
    ("Error Handling", 18, "iface"),
    ("Monitoring", 14, "iface"),
    # Endpoint-level (cols 42-47) — "endpoint" group
    ("Source DB Platform", 18, "endpoint"),
    ("Target DB Platform", 18, "endpoint"),
    ("Source Schema/Object", 20, "endpoint"),
    ("Target Schema/Object", 20, "endpoint"),
    ("Source Tech Platform", 18, "endpoint"),
    ("Target Tech Platform", 18, "endpoint"),
]

FLOWS_HINTS = [
    "e.g. MES Route to ICOST", "1", "e.g. MES 300", "e.g. MES Systems",
    "e.g. XEUS", "e.g. Boundary Apps", "e.g. Direct / API / File", "->",
    "e.g. Near Real-Time", "What data moves", "Why it moves", "Corrections/notes",
    "Tower owner", "Data steward", "e.g. Intel Foundry", "",
    "", "", "", "https://iapm.intel.com/#/app/XXXXX",
    "", "", "", "", "https://iapm.intel.com/#/app/XXXXX",
    # Data arch
    "e.g. Cost Element", "e.g. CSV / IDoc / JSON", "e.g. Intel Confidential",
    "e.g. 10K rows/day", "Master / Transaction", "Lineage notes",
    # Tech arch
    "e.g. Pub-Sub / P2P / ETL", "e.g. Azure Service Bus", "e.g. REST / RFC / SFTP",
    "e.g. OAuth / NTLM / Cert", "DEV,QAS,PRD", "e.g. < 5 min",
    # Interface arch
    "e.g. IF-DS020-001", "e.g. Inbound / Outbound", "e.g. Dead-letter queue",
    "e.g. Splunk / AppDynamics",
    # Endpoint
    "e.g. SAP HANA", "e.g. Azure SQL", "e.g. CKMLHD table",
    "e.g. dbo.CostElements", "e.g. S/4 HANA 2023", "e.g. Azure PaaS",
]

# Tab 2: Business Drivers
BUSINESS_DRIVERS_COLUMNS = [
    ("Driver #", 10), ("Driver Name", 25), ("Description", 45),
    ("Strategic Alignment", 30), ("Priority", 12),
]
BUSINESS_DRIVERS_HINTS = ["1", "e.g. Cost Transparency", "Why this matters...",
                          "e.g. IDM 2.0 Strategy", "High / Medium / Low"]

# Tab 3: Success Criteria
SUCCESS_CRITERIA_COLUMNS = [
    ("Metric", 25), ("Target", 20), ("Measure", 25),
    ("Baseline", 20), ("Owner", 20),
]
SUCCESS_CRITERIA_HINTS = ["e.g. Costing Cycle Time", "< 2 hours",
                          "End-to-end batch runtime", "4 hours (legacy)", "Cost Accounting"]

# Tab 4: Business Architecture (ProcessFlows — BPMN alternative)
BUSINESS_ARCH_COLUMNS = [
    ("Process ID", 14), ("Step #", 8), ("Step Name", 28),
    ("Step Description", 40), ("Step Type", 14),
    ("Lane / Role", 20), ("Source Lane / Role", 18), ("Target Lane / Role", 18),
    ("Gateway Type", 14), ("Gateway Condition", 20),
    ("Preceding Step", 14), ("Following Step", 14),
    ("System / Application", 20), ("Transaction Code", 16), ("Fiori App", 16),
    ("Input Data", 20), ("Output Data", 20), ("Business Rule", 25),
    ("Exception Path", 25), ("SLA / Duration", 14), ("Frequency", 14),
    ("Automation Level", 16), ("Notes", 30),
]
BUSINESS_ARCH_HINTS = [
    "e.g. DS-020-080", "10", "Check Material Price",
    "Verify standard cost matches configuration", "userTask",
    "Cost Accountant", "", "", "none", "",
    "", "20", "SAP S/4 HANA", "CK11N", "F0722",
    "Cost estimate", "Validated price", "Must match within 1% tolerance",
    "Escalate to cost lead", "< 30 min", "Monthly",
    "Semi-Auto", "Runs during month-end close",
]

# Tab 5: NFRs
NFRS_COLUMNS = [
    ("Category", 18), ("Requirement", 40), ("Target / SLA", 20),
    ("Priority", 12), ("Notes", 30),
]
NFRS_HINTS = ["Performance", "Month-end costing batch completes within SLA",
              "< 2 hours end-to-end", "High", "Measured at PRD"]

# Tab 6: Security Controls
SECURITY_COLUMNS = [
    ("Concern", 20), ("Approach", 35), ("Standard / Policy", 25),
    ("Owner", 20), ("Notes", 30),
]
SECURITY_HINTS = ["Authentication", "SSO via Intel corporate identity",
                  "Intel IT Security Policy", "IT Security", ""]

# Tab 7: SAP Dev Status
SAP_STATUS_COLUMNS = [
    ("Object Type", 22), ("Object Name", 25), ("Environment", 14),
    ("Count", 10), ("Status", 14), ("Notes", 30),
]
SAP_STATUS_HINTS = ["Transport Requests", "", "DEV", "42",
                    "Released", "e.g. 42 released TRs"]

# Tab 8: Recommendations
RECOMMENDATIONS_COLUMNS = [
    ("#", 6), ("Category", 18), ("Recommendation", 45),
    ("Priority", 12), ("Owner", 20), ("Target Date", 14), ("Status", 14),
]
RECOMMENDATIONS_HINTS = ["1", "Architecture", "Migrate legacy ICOST flows to S/4 HANA native",
                         "High", "Tower Architect", "2027-Q1", "Open"]

# Master tab registry
TAB_DEFINITIONS = [
    ("Flows", FLOWS_COLUMNS, FLOWS_HINTS),
    ("Business Drivers", BUSINESS_DRIVERS_COLUMNS, BUSINESS_DRIVERS_HINTS),
    ("Success Criteria", SUCCESS_CRITERIA_COLUMNS, SUCCESS_CRITERIA_HINTS),
    ("Business Architecture", BUSINESS_ARCH_COLUMNS, BUSINESS_ARCH_HINTS),
    ("NFRs", NFRS_COLUMNS, NFRS_HINTS),
    ("Security Controls", SECURITY_COLUMNS, SECURITY_HINTS),
    ("SAP Dev Status", SAP_STATUS_COLUMNS, SAP_STATUS_HINTS),
    ("Recommendations", RECOMMENDATIONS_COLUMNS, RECOMMENDATIONS_HINTS),
]


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------
def create_template_workbook(output_path: Path) -> None:
    """Create a multi-tab Excel template workbook with formatted headers and hint rows."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for tab_name, columns, hints in TAB_DEFINITIONS:
        ws = wb.create_sheet(title=tab_name)

        # Write header row
        for col_idx, col_def in enumerate(columns, 1):
            if tab_name == "Flows":
                col_name, width, group = col_def
                fill = GROUP_FILLS.get(group, HEADER_FILL)
            else:
                col_name, width = col_def
                fill = HEADER_FILL

            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = fill
            cell.alignment = HEADER_ALIGN
            cell.border = HEADER_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Write hint row (row 2) — italic gray examples
        for col_idx, hint in enumerate(hints, 1):
            if hint:
                cell = ws.cell(row=2, column=col_idx, value=hint)
                cell.font = HINT_FONT
                cell.fill = HINT_FILL

        # Freeze top row
        ws.freeze_panes = "A2"

        # Auto-filter on header row
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}1"

    # Set Flows as the active tab
    wb.active = 0

    wb.save(str(output_path))


# ---------------------------------------------------------------------------
# CSV → XLSX migration
# ---------------------------------------------------------------------------
def _read_csv_rows(csv_path: Path) -> tuple[list[str], list[list[str]]]:
    """Read a CSV and return (headers, data_rows)."""
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        data = [row for row in reader if any(cell.strip() for cell in row)]
    return headers, data


def migrate_csvs_to_xlsx(data_dir: Path, output_name: str, dry_run: bool = False) -> str:
    """Migrate existing CSV files in a capability's input/data/ into a single xlsx.

    Reads: CurrentFlows.csv (or FutureFlows.csv) + all supplementary CSVs
    Writes: CurrentFlows.xlsx (or FutureFlows.xlsx) with populated tabs

    Returns status string.
    """
    # Determine the flow CSV name
    flow_csv = data_dir / output_name.replace(".xlsx", ".csv")
    if not flow_csv.exists():
        return f"skip-no-{flow_csv.name}"

    output_path = data_dir / output_name
    if output_path.exists():
        return f"exists-{output_name}"

    if dry_run:
        return f"would-create-{output_name}"

    # Start with a template workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Map: CSV filename → tab name
    csv_to_tab = {
        flow_csv.name: "Flows",
        "BusinessDrivers.csv": "Business Drivers",
        "SuccessCriteria.csv": "Success Criteria",
        "ProcessFlows.csv": "Business Architecture",
        "NFRs.csv": "NFRs",
        "SecurityControls.csv": "Security Controls",
        "SAPDevStatus.csv": "SAP Dev Status",
        "Recommendations.csv": "Recommendations",
    }

    tabs_populated = 0

    for tab_name, columns, hints in TAB_DEFINITIONS:
        ws = wb.create_sheet(title=tab_name)

        # Find the matching CSV
        csv_name = None
        for cn, tn in csv_to_tab.items():
            if tn == tab_name:
                csv_name = cn
                break

        csv_path = data_dir / csv_name if csv_name else None

        # Write formatted header row
        for col_idx, col_def in enumerate(columns, 1):
            if tab_name == "Flows":
                col_name, width, group = col_def
                fill = GROUP_FILLS.get(group, HEADER_FILL)
            else:
                col_name, width = col_def
                fill = HEADER_FILL

            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = fill
            cell.alignment = HEADER_ALIGN
            cell.border = HEADER_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # If CSV exists and has data, populate
        if csv_path and csv_path.exists():
            headers, data_rows = _read_csv_rows(csv_path)
            if data_rows:
                # Check if first data row is a hint/example row
                first_non_empty = data_rows[0] if data_rows else []
                is_hint_row = False
                if first_non_empty:
                    # Template hint rows often start with specific patterns
                    first_val = first_non_empty[0].strip().lower() if first_non_empty[0] else ""
                    if first_val in ("", "1") and len(data_rows) == 1:
                        # Likely a template with just placeholder row
                        pass

                for row_idx, data_row in enumerate(data_rows, 2):
                    for col_idx, val in enumerate(data_row, 1):
                        if col_idx <= len(columns) and val.strip():
                            ws.cell(row=row_idx, column=col_idx, value=val.strip())
                tabs_populated += 1
        else:
            # Write hint row for empty tabs
            for col_idx, hint in enumerate(hints, 1):
                if hint:
                    cell = ws.cell(row=2, column=col_idx, value=hint)
                    cell.font = HINT_FONT
                    cell.fill = HINT_FILL

        # Freeze + autofilter
        ws.freeze_panes = "A2"
        last_col = get_column_letter(len(columns))
        ws.auto_filter.ref = f"A1:{last_col}1"

    wb.active = 0
    wb.save(str(output_path))
    return f"created-{tabs_populated}-tabs-populated"


# ---------------------------------------------------------------------------
# Capability discovery
# ---------------------------------------------------------------------------
def _is_cap_id(name: str) -> bool:
    return bool(re.match(r"^[A-Z]+-\d+$", name, re.IGNORECASE))


def discover_capabilities(tower_dir: Path) -> list[Path]:
    caps = []
    for d in sorted(tower_dir.rglob("*")):
        if d.is_dir() and _is_cap_id(d.name) and (d / "input" / "data").is_dir():
            caps.append(d)
    return caps


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Generate and deploy multi-tab Excel input templates")
    parser.add_argument("--deploy", action="store_true", help="Deploy blank templates to capabilities")
    parser.add_argument("--migrate", action="store_true", help="Migrate existing CSVs into xlsx workbooks")
    parser.add_argument("--tower", help="Single tower shortcode (e.g. FPR)")
    parser.add_argument("--cap", help="Single capability ID (e.g. DS-020)")
    parser.add_argument("--dry-run", action="store_true", help="Preview only")
    args = parser.parse_args()

    # Step 1: Always generate the master templates in templates/
    print("Generating master templates...")
    for name in ["CurrentFlows_TEMPLATE.xlsx", "FutureFlows_TEMPLATE.xlsx"]:
        out = TEMPLATES_DIR / name
        if not out.exists() or not args.deploy:
            create_template_workbook(out)
            print(f"  Created: {out.name}")
        else:
            print(f"  Exists:  {out.name}")

    if not args.deploy and not args.migrate:
        print("\nDone. Use --deploy to push to capabilities, --migrate to convert CSVs.")
        return

    # Determine towers
    if args.tower:
        tower_dirs = [TOWERS_DIR / args.tower]
        if not tower_dirs[0].is_dir():
            for d in TOWERS_DIR.iterdir():
                if d.is_dir() and d.name.upper() == args.tower.upper():
                    tower_dirs = [d]
                    break
            else:
                print(f"ERROR: Tower not found: {args.tower}")
                sys.exit(1)
    else:
        tower_dirs = sorted([d for d in TOWERS_DIR.iterdir() if d.is_dir()])

    total_caps = 0
    total_actions = 0

    for tower_dir in tower_dirs:
        print(f"\n{'='*60}")
        print(f"TOWER: {tower_dir.name}")
        print(f"{'='*60}")

        caps = discover_capabilities(tower_dir)
        if args.cap:
            caps = [c for c in caps if c.name == args.cap]

        for cap_dir in caps:
            total_caps += 1
            data_dir = cap_dir / "input" / "data"
            print(f"\n  {cap_dir.name}:")

            if args.migrate:
                # Migrate existing CSVs → xlsx for each flow type present
                for flow_type in ["CurrentFlows", "FutureFlows",
                                  "R3_CurrentFlows", "R3_FutureFlows",
                                  "R4_CurrentFlows", "R4_FutureFlows"]:
                    csv_path = data_dir / f"{flow_type}.csv"
                    if csv_path.exists():
                        status = migrate_csvs_to_xlsx(data_dir, f"{flow_type}.xlsx", args.dry_run)
                        print(f"    {flow_type}.xlsx: {status}")
                        if "created" in status or "would" in status:
                            total_actions += 1

            elif args.deploy:
                # Deploy blank templates
                for flow_type in ["CurrentFlows", "FutureFlows"]:
                    target = data_dir / f"{flow_type}.xlsx"
                    if target.exists():
                        print(f"    {flow_type}.xlsx: exists")
                    elif args.dry_run:
                        print(f"    {flow_type}.xlsx: would-deploy")
                        total_actions += 1
                    else:
                        create_template_workbook(target)
                        print(f"    {flow_type}.xlsx: deployed")
                        total_actions += 1

    mode = "DRY RUN" if args.dry_run else "COMPLETE"
    print(f"\n{'='*60}")
    print(f"{mode}: {total_caps} capabilities, {total_actions} actions")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
