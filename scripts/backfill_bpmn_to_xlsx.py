"""backfill_bpmn_to_xlsx.py — Populate Business Architecture tabs from BPMN XML files.

Reads BPMN 2.0 XML files from each capability's input/bpmn/ directory, extracts
structured process steps (nodes, lanes, gateways, sequence flows), and writes them
into the "Business Architecture" tab of the CurrentFlows.xlsx and FutureFlows.xlsx
workbooks.

This means tower architects don't need to manually fill the Business Architecture
tab when Signavio/BIC BPMN exports are available.

Usage:
    python backfill_bpmn_to_xlsx.py                         # all towers
    python backfill_bpmn_to_xlsx.py --tower FPR              # single tower
    python backfill_bpmn_to_xlsx.py --tower FPR --cap DS-020 # single capability
    python backfill_bpmn_to_xlsx.py --dry-run                # preview only
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure src is importable (scripts/ is one level below project root)
WORKSPACE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(WORKSPACE))
os.chdir(str(WORKSPACE))

from src.bpmn_parser import parse_bpmn, load_capability_bpmns, BPMNProcess, BPMNNode, BPMNFlow

TOWERS_DIR = WORKSPACE / "towers"

# ---------------------------------------------------------------------------
# Styling (matches gen_xlsx_templates.py)
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(
    bottom=Side(style="thin", color="000000"),
    right=Side(style="thin", color="D9D9D9"),
)

# Business Architecture tab columns + widths
BIZ_ARCH_COLUMNS = [
    ("Process ID", 14), ("Step #", 8), ("Step Name", 28),
    ("Step Description", 40), ("Step Type", 14),
    ("Lane / Role", 20), ("Source Lane / Role", 18), ("Target Lane / Role", 18),
    ("Gateway Type", 14), ("Gateway Condition", 20),
    ("Preceding Step", 14), ("Following Step", 14),
    ("System / Application", 20), ("Transaction Code", 16), ("Fiori App", 16),
    ("Input Data", 20), ("Output Data", 20), ("Business Rule", 25),
    ("Exception Path", 25), ("SLA / Duration", 14), ("Frequency", 14),
    ("Automation Level", 16), ("Notes", 30),
]

TAB_NAME = "Business Architecture"


# ---------------------------------------------------------------------------
# BPMN → Row conversion
# ---------------------------------------------------------------------------
def bpmn_to_rows(proc: BPMNProcess) -> list[list[str]]:
    """Convert a BPMNProcess into rows for the Business Architecture tab.

    Each BPMNNode becomes one row. Sequence flows are mapped to
    Preceding Step / Following Step columns.
    """
    if not proc.nodes:
        return []

    # Build adjacency maps from flows
    predecessors: dict[str, list[str]] = {}   # node_id → [source node names]
    successors: dict[str, list[str]] = {}     # node_id → [target node names]
    flow_labels: dict[str, str] = {}          # "src→tgt" → condition label

    # Also build a node ordering based on flow topology
    node_order: dict[str, int] = {}
    ordered_ids = _topological_sort(proc)
    for idx, nid in enumerate(ordered_ids, 1):
        node_order[nid] = idx

    for flow in proc.flows:
        src_node = proc.nodes.get(flow.source_ref)
        tgt_node = proc.nodes.get(flow.target_ref)
        if src_node and tgt_node:
            predecessors.setdefault(flow.target_ref, []).append(
                str(node_order.get(flow.source_ref, "?"))
            )
            successors.setdefault(flow.source_ref, []).append(
                str(node_order.get(flow.target_ref, "?"))
            )
            if flow.name:
                flow_labels[f"{flow.source_ref}→{flow.target_ref}"] = flow.name

    rows = []
    for nid in ordered_ids:
        node = proc.nodes[nid]
        step_num = node_order.get(nid, 0)

        # Map BPMN type to user-friendly Step Type
        type_map = {
            "userTask": "userTask",
            "serviceTask": "serviceTask",
            "task": "task",
            "startEvent": "startEvent",
            "endEvent": "endEvent",
            "intermediateThrowEvent": "intermediateEvent",
            "intermediateCatchEvent": "intermediateEvent",
            "exclusiveGateway": "gateway",
            "parallelGateway": "gateway",
            "inclusiveGateway": "gateway",
            "eventBasedGateway": "gateway",
            "subProcess": "subProcess",
            "callActivity": "subProcess",
            "sendTask": "serviceTask",
            "receiveTask": "serviceTask",
            "scriptTask": "serviceTask",
            "manualTask": "userTask",
            "businessRuleTask": "serviceTask",
        }
        step_type = type_map.get(node.type, "task")

        # Gateway type
        gw_type = ""
        if "exclusive" in node.type.lower():
            gw_type = "exclusive"
        elif "parallel" in node.type.lower():
            gw_type = "parallel"
        elif "inclusive" in node.type.lower():
            gw_type = "inclusive"

        # Gateway conditions from outgoing flows
        gw_conditions = []
        for flow in proc.flows:
            if flow.source_ref == nid and flow.name:
                gw_conditions.append(flow.name.strip())
        gw_condition = " / ".join(gw_conditions)

        # Preceding / Following step numbers
        pred_steps = predecessors.get(nid, [])
        succ_steps = successors.get(nid, [])

        # Determine automation level from type
        auto_level = ""
        if node.type in ("serviceTask", "sendTask", "receiveTask", "scriptTask"):
            auto_level = "Automated"
        elif node.type in ("userTask", "manualTask"):
            auto_level = "Manual"
        elif node.type == "businessRuleTask":
            auto_level = "Semi-Auto"

        row = [
            proc.step_id,                           # Process ID
            str(step_num),                           # Step #
            node.display_name,                       # Step Name
            "",                                      # Step Description (not in BPMN)
            step_type,                               # Step Type
            node.lane,                               # Lane / Role
            "",                                      # Source Lane / Role
            "",                                      # Target Lane / Role
            gw_type,                                 # Gateway Type
            gw_condition,                            # Gateway Condition
            ", ".join(pred_steps),                    # Preceding Step
            ", ".join(succ_steps),                    # Following Step
            "",                                      # System / Application
            "",                                      # Transaction Code
            "",                                      # Fiori App
            "",                                      # Input Data
            "",                                      # Output Data
            "",                                      # Business Rule
            "",                                      # Exception Path
            "",                                      # SLA / Duration
            "",                                      # Frequency
            auto_level,                              # Automation Level
            "",                                      # Notes
        ]
        rows.append(row)

    return rows


def _topological_sort(proc: BPMNProcess) -> list[str]:
    """Order nodes by sequence flow topology (BFS from start events).

    Falls back to insertion order if no start event exists.
    """
    # Find start events
    start_ids = [nid for nid, n in proc.nodes.items() if n.type == "startEvent"]
    if not start_ids:
        return list(proc.nodes.keys())

    # Build adjacency list
    adj: dict[str, list[str]] = {}
    for flow in proc.flows:
        adj.setdefault(flow.source_ref, []).append(flow.target_ref)

    # BFS
    visited: set[str] = set()
    order: list[str] = []
    queue = list(start_ids)

    while queue:
        nid = queue.pop(0)
        if nid in visited or nid not in proc.nodes:
            continue
        visited.add(nid)
        order.append(nid)
        for next_id in adj.get(nid, []):
            if next_id not in visited:
                queue.append(next_id)

    # Add any unreachable nodes at the end
    for nid in proc.nodes:
        if nid not in visited:
            order.append(nid)

    return order


# ---------------------------------------------------------------------------
# Write to xlsx
# ---------------------------------------------------------------------------
def write_biz_arch_tab(xlsx_path: Path, processes: list[BPMNProcess], dry_run: bool = False) -> str:
    """Write/overwrite the Business Architecture tab in an xlsx workbook.

    Returns status string.
    """
    if not xlsx_path.exists():
        return "skip-no-xlsx"

    if not processes:
        return "skip-no-bpmn"

    # Collect all rows from all processes
    all_rows = []
    for proc in processes:
        rows = bpmn_to_rows(proc)
        all_rows.extend(rows)

    if not all_rows:
        return "skip-empty-bpmn"

    if dry_run:
        return f"would-write-{len(all_rows)}-steps-from-{len(processes)}-processes"

    # Open workbook
    wb = openpyxl.load_workbook(str(xlsx_path))

    # Remove existing Business Architecture sheet if present
    if TAB_NAME in wb.sheetnames:
        del wb[TAB_NAME]

    # Create new sheet at position 3 (after Flows, Business Drivers, Success Criteria)
    ws = wb.create_sheet(title=TAB_NAME, index=3)

    # Write header
    for col_idx, (col_name, width) in enumerate(BIZ_ARCH_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = HEADER_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data rows
    for row_idx, row_data in enumerate(all_rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            if val:
                ws.cell(row=row_idx, column=col_idx, value=val)

    # Freeze + autofilter
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(BIZ_ARCH_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}1"

    wb.save(str(xlsx_path))
    return f"wrote-{len(all_rows)}-steps-from-{len(processes)}-processes"


# ---------------------------------------------------------------------------
# Capability discovery
# ---------------------------------------------------------------------------
def _is_cap_id(name: str) -> bool:
    return bool(re.match(r"^[A-Z]+-\d+$", name, re.IGNORECASE))


def discover_capabilities(tower_dir: Path) -> list[Path]:
    caps = []
    for root, dirs, files in os.walk(str(tower_dir)):
        d = Path(root)
        if d.is_dir() and _is_cap_id(d.name) and (d / "input" / "data").is_dir():
            caps.append(d)
    return sorted(caps, key=lambda p: p.name)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Backfill Business Architecture tabs from BPMN files")
    parser.add_argument("--tower", help="Single tower shortcode")
    parser.add_argument("--cap", help="Single capability ID")
    parser.add_argument("--dry-run", action="store_true", help="Preview only")
    args = parser.parse_args()

    if args.tower:
        tower_dirs = [TOWERS_DIR / args.tower]
        if not tower_dirs[0].is_dir():
            for d in TOWERS_DIR.iterdir():
                if d.is_dir() and d.name.upper() == args.tower.upper():
                    tower_dirs = [d]
                    break
            else:
                print(f"ERROR: Tower not found: {args.tower}")
                sys.exit(1)
    else:
        tower_dirs = sorted([d for d in TOWERS_DIR.iterdir() if d.is_dir()])

    total_caps = 0
    total_populated = 0
    total_steps = 0
    total_skipped = 0

    for tower_dir in tower_dirs:
        print(f"\n{'='*60}")
        print(f"TOWER: {tower_dir.name}")
        print(f"{'='*60}")

        caps = discover_capabilities(tower_dir)
        if args.cap:
            caps = [c for c in caps if c.name == args.cap]

        for cap_dir in caps:
            total_caps += 1
            bpmn_dir = cap_dir / "input" / "bpmn"
            data_dir = cap_dir / "input" / "data"

            # Load BPMN processes
            processes = load_capability_bpmns(bpmn_dir)
            if not processes:
                total_skipped += 1
                continue

            print(f"\n  {cap_dir.name}: {len(processes)} BPMN processes")

            # Write to each xlsx in data_dir
            for xlsx_file in sorted(data_dir.glob("*.xlsx")):
                status = write_biz_arch_tab(xlsx_file, processes, args.dry_run)
                print(f"    {xlsx_file.name}: {status}")
                if "wrote" in status or "would-write" in status:
                    total_populated += 1
                    # Extract step count from status
                    parts = status.split("-")
                    if len(parts) >= 2 and parts[1].isdigit():
                        total_steps += int(parts[1])

    mode = "DRY RUN" if args.dry_run else "COMPLETE"
    print(f"\n{'='*60}")
    print(f"BACKFILL {mode}")
    print(f"  Capabilities processed: {total_caps}")
    print(f"  Capabilities with BPMN: {total_caps - total_skipped}")
    print(f"  Workbooks populated:    {total_populated}")
    print(f"  Total process steps:    {total_steps}")
    print(f"  Capabilities skipped:   {total_skipped}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
